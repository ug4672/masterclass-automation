#!/usr/bin/env python3
"""Local proxy server — serves the HTML and forwards API calls to Jira/Slack/BigQuery."""
import http.server, urllib.request, urllib.parse, urllib.error, json, os, re, ssl, decimal, datetime
import threading, time, hmac, hashlib, zoneinfo
from concurrent.futures import ThreadPoolExecutor

GCS_BUCKET = os.environ.get('GCS_BUCKET', '')
GCS_CONFIG_KEY = 'snapshot_config.json'

# Our own GCP project (separate from ik-marketing-data which holds source data).
# Hosts events.event + history.event_snapshot.
BQ_APP_PROJECT = os.environ.get('BQ_APP_PROJECT', 'masterclass-automation-ik')

# macOS Python from python.org often lacks system certs — use unverified context for localhost proxy
SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE

PORT = int(os.environ.get('PORT', 8080))

SESSION_SECRET  = os.environ.get('SESSION_SECRET', 'dev-only-change-in-prod')
ALLOWED_DOMAIN  = 'interviewkickstart.com'
GOOGLE_CLIENT_ID = os.environ.get('GOOGLE_CLIENT_ID', '')
SESSION_TTL     = 604800  # 7 days

def _make_session(email):
    ts  = str(int(time.time()))
    msg = email + '|' + ts
    sig = hmac.new(SESSION_SECRET.encode(), msg.encode(), hashlib.sha256).hexdigest()
    return email + '|' + ts + '|' + sig

def _verify_session(token):
    try:
        parts = token.split('|')
        if len(parts) != 3:
            return None
        email, ts, sig = parts
        msg      = email + '|' + ts
        expected = hmac.new(SESSION_SECRET.encode(), msg.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(sig, expected):
            return None
        if int(time.time()) - int(ts) > SESSION_TTL:
            return None
        return email
    except Exception:
        return None

def _session_email(handler):
    for part in handler.headers.get('Cookie', '').split(';'):
        k, _, v = part.strip().partition('=')
        if k.strip() == 'ik_session':
            return _verify_session(v.strip())
    return None

def json_serial(obj):
    if isinstance(obj, (datetime.date, datetime.datetime)):
        return obj.isoformat()
    if isinstance(obj, decimal.Decimal):
        return float(obj)
    return str(obj)

def _row_to_dict(r):
    """BigQuery Row → JSON-safe dict. Timestamps → ISO strings, decimals → floats."""
    out = {}
    for k in r.keys():
        v = r.get(k)
        if isinstance(v, (datetime.date, datetime.datetime)):
            out[k] = v.isoformat()
        elif isinstance(v, decimal.Decimal):
            out[k] = float(v)
        else:
            out[k] = v
    return out

def _cell_to_str(v):
    """Format a BQ value for a Google Sheets cell."""
    if v is None:
        return ''
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.isoformat()
    if isinstance(v, decimal.Decimal):
        return float(v)
    if isinstance(v, (dict, list)):
        return json.dumps(v, default=str)
    return v if isinstance(v, (int, float, bool)) else str(v)

def _generate_event_id(body):
    """Slug: <first-4-title-words>-<instructor-initials>-<mmm-dd>."""
    title      = (body.get('title') or '').strip()
    instructor = (body.get('instructorName') or '').strip()
    live_at    = body.get('liveAt') or ''

    words      = re.findall(r'[a-z0-9]+', title.lower())
    title_part = '-'.join(words[:4])[:40] or 'event'
    initials   = ''.join(w[0] for w in instructor.split() if w)[:3].lower()

    date_part = ''
    if live_at:
        try:
            dt = datetime.datetime.fromisoformat(live_at.replace('Z', '+00:00'))
            date_part = dt.strftime('%b%d').lower()
        except Exception:
            pass

    parts = [title_part]
    if initials:  parts.append(initials)
    if date_part: parts.append(date_part)
    return '-'.join(parts)

class Handler(http.server.SimpleHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self._cors()
        self.end_headers()

    def do_HEAD(self):
        if urllib.parse.urlparse(self.path).path == '/':
            self.send_response(302)
            self.send_header('Location', '/hub.html')
            self.end_headers()
            return
        super().do_HEAD()

    def do_GET(self):
        p = urllib.parse.urlparse(self.path)
        path = p.path
        if path == '/':
            self.send_response(302)
            self.send_header('Location', '/hub.html')
            self.end_headers()
            return
        if path == '/auth/check':
            email = _session_email(self)
            if email:
                self._json_response(200, {'ok': True, 'email': email})
            else:
                self._json_response(401, {'error': 'Not authenticated'})
            return
        if path == '/events' or path.startswith('/events/'):
            if not _session_email(self):
                self._json_response(401, {'error': 'Not authenticated'})
                return
            if path == '/events':
                self._list_events(p)
            else:
                self._get_event(path[len('/events/'):])
            return
        if path == '/months':
            if not _session_email(self):
                self._json_response(401, {'error': 'Not authenticated'})
                return
            self._list_months(p)
            return
        if path == '/series':
            if not _session_email(self):
                self._json_response(401, {'error': 'Not authenticated'})
                return
            self._list_series(p)
            return
        if path == '/sales':
            if not _session_email(self):
                self._json_response(401, {'error': 'Not authenticated'})
                return
            self._list_sales(p)
            return
        if path == '/event/poll-qna-export':
            if not _session_email(self):
                self._json_response(401, {'error': 'Not authenticated'})
                return
            self._poll_qna_export(p)
            return
        super().do_GET()

    def do_POST(self):
        # Public endpoints — no session required
        if self.path == '/auth/verify':
            self._auth_verify()
            return
        if self.path == '/auth/logout':
            self._auth_logout()
            return
        if self.path.startswith('/run-snapshot'):  # called by Cloud Scheduler, no browser session
            self._run_snapshot()
            return
        # All other endpoints require a valid session
        if not _session_email(self):
            self._json_response(401, {'error': 'Not authenticated'})
            return
        if self.path.startswith('/proxy'):
            self._proxy()
        elif self.path == '/bigquery':
            self._bigquery()
        elif self.path == '/save-snapshot-config':
            self._save_snapshot_config()
        elif self.path == '/events':
            self._create_event()
        elif self.path.startswith('/events/'):
            # Per-event refresh: POST /events/<id>/refresh
            p = urllib.parse.urlparse(self.path)
            if p.path.endswith('/refresh'):
                eid = urllib.parse.unquote(p.path[len('/events/'):-len('/refresh')])
                self._refresh_event(eid)
            else:
                self.send_response(404)
                self.end_headers()
        else:
            self.send_response(404)
            self.end_headers()

    def _cors(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Authorization, Content-Type, Accept')

    def _json_response(self, status, data, extra_headers=None):
        body = json.dumps(data, default=json_serial).encode()
        self.send_response(status)
        self._cors()
        self.send_header('Content-Type', 'application/json')
        for k, v in (extra_headers or {}).items():
            self.send_header(k, v)
        self.end_headers()
        self.wfile.write(body)

    def _proxy(self):
        params = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
        target = params.get('url', [''])[0]
        if not target:
            self.send_response(400)
            self.end_headers()
            return

        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length) if length else None

        headers = {}
        for h in ['Authorization', 'Content-Type', 'Accept']:
            if self.headers.get(h):
                headers[h] = self.headers[h]

        try:
            req = urllib.request.Request(target, data=body, headers=headers, method='POST')
            with urllib.request.urlopen(req, context=SSL_CTX) as r:
                data = r.read()
                self.send_response(r.status)
                self._cors()
                self.send_header('Content-Type', r.headers.get('Content-Type', 'application/json'))
                self.end_headers()
                self.wfile.write(data)
        except urllib.error.HTTPError as e:
            data = e.read()
            self.send_response(e.code)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(data)
        except Exception as e:
            self._json_response(500, {'error': str(e)})

    def _bigquery(self):
        length = int(self.headers.get('Content-Length', 0))
        body = json.loads(self.rfile.read(length)) if length else {}
        query    = body.get('query', '')
        key_file = body.get('keyFile', '').strip()
        project  = body.get('project', '').strip() or None

        try:
            from google.cloud import bigquery
            from google.oauth2 import service_account

            if key_file:
                creds  = service_account.Credentials.from_service_account_file(key_file)
                client = bigquery.Client(credentials=creds, project=project or creds.project_id)
            else:
                client = bigquery.Client(project=project)

            job    = client.query(query)
            result = job.result()
            schema = [f.name for f in result.schema]
            rows   = [list(row.values()) for row in result]
            self._json_response(200, {'schema': schema, 'rows': rows})

        except ImportError:
            self._json_response(500, {'error': 'google-cloud-bigquery not installed. Run: pip install google-cloud-bigquery'})
        except Exception as e:
            self._json_response(500, {'error': str(e)})

    def _list_events(self, parsed_url):
        from google.cloud import bigquery
        qs      = urllib.parse.parse_qs(parsed_url.query)
        status  = qs.get('status', ['upcoming'])[0]
        country = qs.get('country', [None])[0]
        ids_csv = qs.get('ids', [None])[0]
        try:
            limit = min(int(qs.get('limit', ['50'])[0]), 1000)
        except ValueError:
            limit = 50

        where  = ["COALESCE(e.status, 'upcoming') != 'archived'"]
        params = []
        if ids_csv:
            # When ids are supplied, status/country filters are ignored — the caller knows
            # what it wants (typically the Compare workspace passing a deep-link).
            id_list = [i for i in ids_csv.split(',') if i]
            where = ["e.event_id IN UNNEST(@ids)"]
            params.append(bigquery.ArrayQueryParameter('ids', 'STRING', id_list))
            limit = max(limit, len(id_list))
        else:
            if status == 'upcoming':
                where.append("e.live_at > TIMESTAMP_SUB(CURRENT_TIMESTAMP(), INTERVAL 4 HOUR)")
            elif status == 'aired':
                where.append("e.live_at <= TIMESTAMP_SUB(CURRENT_TIMESTAMP(), INTERVAL 4 HOUR)")
            if country:
                where.append("LOWER(e.country) = LOWER(@country)")
                params.append(bigquery.ScalarQueryParameter('country', 'STRING', country))

        order = 'ASC' if status == 'upcoming' else 'DESC'
        query = f"""
WITH latest_snap AS (
  SELECT *
  FROM `{BQ_APP_PROJECT}.history.event_snapshot`
  QUALIFY ROW_NUMBER() OVER (PARTITION BY event_id ORDER BY snapshot_at DESC) = 1
)
SELECT
  e.event_id, e.title, e.topic, e.event_type, e.country, e.webinar_type, e.series,
  e.live_at, e.day2_live_at, e.go_live_date, e.landing_url, e.slides_url, e.yt_url, e.zoom_url,
  e.instructor_name, e.instructor_role, e.goal_regs, e.status,
  s.total_regs, s.meta_regs, s.crm_regs, s.other_regs,
  s.meta_spend, s.cpiql, s.attendees, s.attendance_pct,
  s.hours_to_live, s.snapshot_at,
  s.email_sent, s.email_delivered, s.email_opened, s.email_clicked,
  s.calls_attempted, s.calls_connected, s.avg_talk_seconds,
  s.role_sde, s.role_ml, s.role_management, s.role_systems, s.role_null, s.role_other,
  s.we_3_5, s.we_6_10, s.we_10_15, s.we_15_20, s.we_20p, s.we_other,
  s.us_yt_regs, s.us_social_regs, s.us_l10x_email_regs, s.us_l10x_bot_regs,
  s.us_ni_base_regs, s.us_other_regs,
  s.extras,
  s.call_total_leads,
  s.call_pre_attempts, s.call_pre_connects, s.call_pre_talk_mins, s.call_pre_covered,
  s.call_p2_attempts,  s.call_p2_connects,  s.call_p2_talk_mins,  s.call_p2_covered,
  s.call_p7_attempts,  s.call_p7_connects,  s.call_p7_talk_mins,  s.call_p7_covered,
  s.call_p14_attempts, s.call_p14_connects, s.call_p14_talk_mins, s.call_p14_covered,
  s.call_p14p_attempts,s.call_p14p_connects,s.call_p14p_talk_mins,s.call_p14p_covered,
  s.sales, s.revenue, s.paid_revenue, s.overall_roas, s.paid_roas
FROM `{BQ_APP_PROJECT}.events.event` e
LEFT JOIN latest_snap s USING (event_id)
WHERE {' AND '.join(where)}
ORDER BY e.live_at {order}
LIMIT {limit}"""

        try:
            client = bigquery.Client(project=BQ_APP_PROJECT)
            rows   = list(client.query(query, job_config=bigquery.QueryJobConfig(query_parameters=params)).result())
        except Exception as e:
            self._json_response(500, {'error': str(e)})
            return

        events = []
        snap_keys = ('total_regs', 'meta_regs', 'crm_regs', 'other_regs', 'meta_spend',
                     'cpiql', 'attendees', 'attendance_pct', 'hours_to_live', 'snapshot_at',
                     'email_sent', 'email_delivered', 'email_opened', 'email_clicked',
                     'calls_attempted', 'calls_connected', 'avg_talk_seconds',
                     'role_sde', 'role_ml', 'role_management', 'role_systems', 'role_null', 'role_other',
                     'we_3_5', 'we_6_10', 'we_10_15', 'we_15_20', 'we_20p', 'we_other',
                     'us_yt_regs', 'us_social_regs', 'us_l10x_email_regs', 'us_l10x_bot_regs',
                     'us_ni_base_regs', 'us_other_regs', 'extras',
                     'call_total_leads',
                     'call_pre_attempts', 'call_pre_connects', 'call_pre_talk_mins', 'call_pre_covered',
                     'call_p2_attempts',  'call_p2_connects',  'call_p2_talk_mins',  'call_p2_covered',
                     'call_p7_attempts',  'call_p7_connects',  'call_p7_talk_mins',  'call_p7_covered',
                     'call_p14_attempts', 'call_p14_connects', 'call_p14_talk_mins', 'call_p14_covered',
                     'call_p14p_attempts','call_p14p_connects','call_p14p_talk_mins','call_p14p_covered',
                     'sales', 'revenue', 'paid_revenue', 'overall_roas', 'paid_roas')
        for r in rows:
            d = _row_to_dict(r)
            ev = {k: v for k, v in d.items() if k not in snap_keys}
            ev['snapshot'] = {k: d.get(k) for k in snap_keys}
            events.append(ev)
        self._json_response(200, {'events': events})

    def _list_series(self, parsed_url):
        from google.cloud import bigquery
        qs      = urllib.parse.parse_qs(parsed_url.query)
        country = qs.get('country', ['India'])[0]
        params  = [bigquery.ScalarQueryParameter('country', 'STRING', country)]
        query = f"""
SELECT series, COUNT(*) AS n
FROM `{BQ_APP_PROJECT}.events.event`
WHERE series IS NOT NULL
  AND LOWER(country) = LOWER(@country)
  AND COALESCE(status, 'upcoming') != 'archived'
GROUP BY series
ORDER BY MAX(live_at) DESC"""
        try:
            client = bigquery.Client(project=BQ_APP_PROJECT)
            rows   = list(client.query(query, job_config=bigquery.QueryJobConfig(query_parameters=params)).result())
        except Exception as e:
            self._json_response(500, {'error': str(e)})
            return
        self._json_response(200, {'series': [{'name': r['series'], 'count': int(r['n'])} for r in rows]},
                            extra_headers={'Cache-Control': 'private, max-age=120'})

    def _list_months(self, parsed_url):
        """Aggregate per-event-month for one country, from Jan 2026 onward.
        Returns one row per month with all metrics (Acquisition + Engagement + Quality +
        Calls), plus a nested list of per-event rows for inline drill-down. The frontend
        picks which metric tab to surface.

        Month bucket = event's local-timezone live_at (IST for India, PT for US) — NOT
        lead registration date.

        Blended CPL = sum(spend) / sum(paid_regs) across the month, which is more honest
        than averaging per-event CPLs of wildly different volumes."""
        from google.cloud import bigquery
        qs      = urllib.parse.parse_qs(parsed_url.query)
        country = qs.get('country', ['India'])[0]
        series  = qs.get('series', [None])[0]
        tz      = 'America/Los_Angeles' if country.lower() in ('us', 'usa') else 'Asia/Kolkata'
        params  = [bigquery.ScalarQueryParameter('country', 'STRING', country)]
        where   = ["LOWER(e.country) = LOWER(@country)",
                   "e.live_at >= TIMESTAMP '2026-01-01 00:00:00'",
                   "COALESCE(e.status, 'upcoming') != 'archived'"]
        if series:
            where.append("e.series = @series")
            params.append(bigquery.ScalarQueryParameter('series', 'STRING', series))

        where_clause = ' AND '.join(where)
        events_query = f"""
WITH latest AS (
  SELECT * FROM `{BQ_APP_PROJECT}.history.event_snapshot`
  QUALIFY ROW_NUMBER() OVER (PARTITION BY event_id ORDER BY snapshot_at DESC) = 1
)
SELECT
  FORMAT_DATE('%Y-%m', DATE_TRUNC(DATE(e.live_at, '{tz}'), MONTH)) AS month_slug,
  e.event_id, e.title, e.instructor_name, e.series, e.live_at,
  s.total_regs, s.meta_regs, s.crm_regs, s.other_regs, s.meta_spend, s.cpiql,
  s.attendees, s.attendance_pct,
  s.email_sent, s.email_delivered, s.email_opened, s.email_clicked,
  s.calls_attempted, s.calls_connected, s.avg_talk_seconds,
  s.call_total_leads, s.call_pre_covered,
  s.call_pre_attempts, s.call_pre_connects, s.call_pre_talk_mins,
  s.call_p2_attempts,  s.call_p2_connects,  s.call_p2_talk_mins,  s.call_p2_covered,
  s.call_p7_attempts,  s.call_p7_connects,  s.call_p7_talk_mins,  s.call_p7_covered,
  s.call_p14_attempts, s.call_p14_connects, s.call_p14_talk_mins, s.call_p14_covered,
  s.call_p14p_attempts,s.call_p14p_connects,s.call_p14p_talk_mins,s.call_p14p_covered,
  s.role_sde, s.role_ml, s.role_management, s.role_systems, s.role_null, s.role_other,
  s.we_3_5, s.we_6_10, s.we_10_15, s.we_15_20, s.we_20p, s.we_other,
  s.us_yt_regs, s.us_social_regs, s.us_l10x_email_regs, s.us_l10x_bot_regs,
  s.us_ni_base_regs, s.us_other_regs,
  s.sales, s.revenue, s.paid_revenue, s.overall_roas, s.paid_roas
FROM `{BQ_APP_PROJECT}.events.event` e
LEFT JOIN latest s USING (event_id)
WHERE {where_clause}
ORDER BY e.live_at DESC"""
        try:
            client = bigquery.Client(project=BQ_APP_PROJECT)
            rows   = list(client.query(events_query, job_config=bigquery.QueryJobConfig(query_parameters=params)).result())
        except Exception as e:
            self._json_response(500, {'error': str(e)})
            return

        events = [_row_to_dict(r) for r in rows]
        # Aggregate to monthly rows in Python — simpler than two BQ queries, fast for
        # ~30 events × ~12 months. Reuses the same row data both for the headline
        # table and for the per-event drill-down (frontend just groups by month_slug).
        by_month = {}
        for ev in events:
            ms = ev.get('month_slug')
            if not ms:
                continue
            m = by_month.setdefault(ms, {
                'month_slug': ms, 'events': 0,
                'total_regs': 0, 'meta_regs': 0, 'crm_regs': 0, 'other_regs': 0,
                'meta_spend': 0.0, 'attendees': 0,
                'email_sent': 0, 'email_delivered': 0, 'email_opened': 0, 'email_clicked': 0,
                'calls_attempted': 0, 'calls_connected': 0, 'talk_seconds_total': 0.0,
                'call_total_leads': 0, 'call_pre_covered': 0,
                'call_pre_attempts': 0, 'call_pre_connects': 0, 'call_pre_talk_mins': 0.0,
                'call_p2_attempts':  0, 'call_p2_connects':  0, 'call_p2_talk_mins':  0.0, 'call_p2_covered':  0,
                'call_p7_attempts':  0, 'call_p7_connects':  0, 'call_p7_talk_mins':  0.0, 'call_p7_covered':  0,
                'call_p14_attempts': 0, 'call_p14_connects': 0, 'call_p14_talk_mins': 0.0, 'call_p14_covered': 0,
                'call_p14p_attempts':0, 'call_p14p_connects':0, 'call_p14p_talk_mins':0.0, 'call_p14p_covered':0,
                'role_sde': 0, 'role_ml': 0, 'role_management': 0, 'role_systems': 0, 'role_null': 0, 'role_other': 0,
                'we_3_5': 0, 'we_6_10': 0, 'we_10_15': 0, 'we_15_20': 0, 'we_20p': 0, 'we_other': 0,
                'us_yt_regs': 0, 'us_social_regs': 0, 'us_l10x_email_regs': 0,
                'us_l10x_bot_regs': 0, 'us_ni_base_regs': 0, 'us_other_regs': 0,
                'sales': 0, 'revenue': 0.0, 'paid_revenue': 0.0,
                # Sale-date view (booked in this calendar month, regardless of MC date)
                'sales_booked': 0, 'revenue_booked': 0.0,
                'event_ids': [],
            })
            m['events'] += 1
            m['event_ids'].append(ev['event_id'])
            sum_keys = ('total_regs', 'meta_regs', 'crm_regs', 'other_regs', 'meta_spend',
                        'attendees', 'email_sent', 'email_delivered', 'email_opened', 'email_clicked',
                        'calls_attempted', 'calls_connected', 'call_total_leads', 'call_pre_covered',
                        'call_pre_attempts', 'call_pre_connects', 'call_pre_talk_mins',
                        'call_p2_attempts',  'call_p2_connects',  'call_p2_talk_mins',  'call_p2_covered',
                        'call_p7_attempts',  'call_p7_connects',  'call_p7_talk_mins',  'call_p7_covered',
                        'call_p14_attempts', 'call_p14_connects', 'call_p14_talk_mins', 'call_p14_covered',
                        'call_p14p_attempts','call_p14p_connects','call_p14p_talk_mins','call_p14p_covered',
                        'role_sde', 'role_ml', 'role_management', 'role_systems', 'role_null', 'role_other',
                        'we_3_5', 'we_6_10', 'we_10_15', 'we_15_20', 'we_20p', 'we_other',
                        'us_yt_regs', 'us_social_regs', 'us_l10x_email_regs',
                        'us_l10x_bot_regs', 'us_ni_base_regs', 'us_other_regs',
                        'sales', 'revenue', 'paid_revenue')
            float_keys = {'meta_spend', 'revenue', 'paid_revenue',
                          'call_pre_talk_mins', 'call_p2_talk_mins', 'call_p7_talk_mins',
                          'call_p14_talk_mins', 'call_p14p_talk_mins'}
            for k in sum_keys:
                v = ev.get(k)
                if v is not None:
                    m[k] += float(v) if k in float_keys else int(v)
            # Talk seconds = avg_talk × connected (weighted reconstruction)
            if ev.get('avg_talk_seconds') is not None and ev.get('calls_connected') is not None:
                m['talk_seconds_total'] += float(ev['avg_talk_seconds']) * int(ev['calls_connected'])

        # Derived ratios per month
        for m in by_month.values():
            m['cpl_blended']      = (m['meta_spend'] / m['meta_regs']) if m['meta_regs'] > 0 else None
            m['attendance_pct']   = (100.0 * m['attendees'] / m['total_regs']) if m['total_regs'] > 0 else None
            m['email_open_pct']   = (100.0 * m['email_opened']   / m['email_sent']) if m['email_sent']   > 0 else None
            m['email_click_pct']  = (100.0 * m['email_clicked']  / m['email_sent']) if m['email_sent']   > 0 else None
            m['call_connect_pct'] = (100.0 * m['calls_connected'] / m['calls_attempted']) if m['calls_attempted'] > 0 else None
            m['avg_talk_seconds'] = (m['talk_seconds_total'] / m['calls_connected']) if m['calls_connected'] > 0 else None
            m['calls_per_lead']   = (m['calls_attempted'] / m['total_regs']) if m['total_regs'] > 0 else None
            # Coverage = leads with ≥1 pre-webinar call attempt / total registrations
            # (pre_covered = COUNTIF(pre_calls > 0) at event level, summed across events)
            m['call_coverage_pct']            = (100.0 * m['call_pre_covered'] / m['total_regs']) if m['total_regs'] > 0 else None
            # Per-user efficiency: avg connected calls per registered lead
            m['avg_calls_connected_per_lead'] = (m['calls_connected'] / m['total_regs']) if m['total_regs'] > 0 else None
            # Per-phase derived metrics (calls/user, coverage%, connected/user, avg talk per connect)
            # Phases: pre, 0–2D, 0–7D, 0–14D, 14D+ (the cumulative windows match BigQuery)
            phases = {}
            for ph in ('pre', 'p2', 'p7', 'p14', 'p14p'):
                att     = m.get(f'call_{ph}_attempts') or 0
                conn    = m.get(f'call_{ph}_connects') or 0
                talkmin = m.get(f'call_{ph}_talk_mins') or 0.0
                covered = m.get(f'call_{ph}_covered') if ph != 'pre' else m.get('call_pre_covered')
                covered = covered or 0
                phases[ph] = {
                    'cpu':  (att / m['total_regs'])              if m['total_regs'] > 0 else None,
                    'cov':  (100.0 * covered / m['total_regs'])  if m['total_regs'] > 0 else None,
                    'conn': (conn / m['total_regs'])             if m['total_regs'] > 0 else None,
                    'talk': (talkmin * 60.0 / conn)              if conn > 0 else None,
                }
            m['call_phases'] = phases
            m['overall_roas']     = (m['revenue'] / m['meta_spend']) if m['meta_spend'] > 0 else None
            m['paid_roas']        = (m['paid_revenue'] / m['meta_spend']) if m['meta_spend'] > 0 else None
            # Defaults; populated below for India only.
            m['crm_email_regs']    = None
            m['crm_whatsapp_regs'] = None

        # India-only: split CRM into Email vs WhatsApp by utm_campaign.
        # WhatsApp = LOWER(utm_campaign) LIKE '%whatsapp%' ; Email = the rest of CRM.
        if country.lower() in ('india', 'in'):
            try:
                src_client = bigquery.Client(project='ik-marketing-data')
                series_pred = ''
                series_params = []
                if series:
                    series_pred = ' AND e.series = @series'
                    series_params = [bigquery.ScalarQueryParameter('series', 'STRING', series)]
                crm_query = f"""
WITH event_air AS (
  SELECT event_id, webinar_type,
    DATE(live_at, "Asia/Kolkata") AS air_pt,
    FORMAT_DATE('%Y-%m', DATE_TRUNC(DATE(live_at, "Asia/Kolkata"), MONTH)) AS month_slug
  FROM `{BQ_APP_PROJECT}.events.event` e
  WHERE LOWER(e.country) = 'india'
    AND e.live_at >= TIMESTAMP '2026-01-01 00:00:00'
    AND COALESCE(e.status, 'upcoming') != 'archived'
    {series_pred}
),
leads_raw AS (
  SELECT hubspot_ID, utm_campaign, Channel, formatted_date,
    DATE(event_start_date_time, "Asia/Kolkata") AS event_pt,
    webinar_type, dupe_flag, gql_flag, work_ex,
    ROW_NUMBER() OVER (
      PARTITION BY hubspot_ID, DATE(event_start_date_time, "Asia/Kolkata")
      ORDER BY formatted_date ASC
    ) AS rnk
  FROM `ik-marketing-data.India_Leads.US_Domain_combined_view`
  WHERE dupe_logic = 1
),
qualified AS (
  SELECT *
  FROM leads_raw
  WHERE rnk = 1 AND dupe_flag = 0 AND gql_flag = 0
    AND LOWER(work_ex) NOT LIKE '%student%' AND work_ex NOT IN ('0-2','3-4')
    AND (REGEXP_CONTAINS(LOWER(Channel), r'email|whatsapp|whats.?app')
         OR LOWER(utm_campaign) LIKE '%whatsapp%')
)
SELECT e.month_slug,
       CASE WHEN LOWER(q.utm_campaign) LIKE '%whatsapp%' THEN 'whatsapp' ELSE 'email' END AS bucket,
       COUNT(*) AS n
FROM event_air e
JOIN qualified q
  ON q.webinar_type = e.webinar_type
  AND q.event_pt    = e.air_pt
GROUP BY month_slug, bucket"""
                rows = list(src_client.query(crm_query, job_config=bigquery.QueryJobConfig(query_parameters=series_params)).result())
                # Merge counts into by_month
                for r in rows:
                    ms = r['month_slug']
                    if ms not in by_month: continue
                    if by_month[ms]['crm_email_regs'] is None:
                        by_month[ms]['crm_email_regs'] = 0
                    if by_month[ms]['crm_whatsapp_regs'] is None:
                        by_month[ms]['crm_whatsapp_regs'] = 0
                    if r['bucket'] == 'whatsapp':
                        by_month[ms]['crm_whatsapp_regs'] = int(r['n'])
                    else:
                        by_month[ms]['crm_email_regs'] = int(r['n'])
            except Exception as e:
                print(f'[months/crm_breakdown] {e}')

        # ── Booked-revenue view: sales grouped by Sale_date month (not MC month) ──
        # Same sales scope as _query_sales_* (tied to a tracked event, alumni-excluded,
        # dedup rules), but bucketed by when the money came in rather than when the
        # masterclass was held. Populates revenue_booked / sales_booked per month.
        try:
            src_client = bigquery.Client(project='ik-marketing-data')
            if country.lower() in ('india', 'in'):
                booked_query = f"""
WITH events_tracked AS (
  SELECT UPPER(webinar_type) AS wt, DATE(live_at, 'Asia/Kolkata') AS web_dt
  FROM `{BQ_APP_PROJECT}.events.event`
  WHERE LOWER(country) = 'india' AND live_at >= TIMESTAMP '2026-01-01 00:00:00'
    AND COALESCE(status, 'upcoming') != 'archived'
),
sales_raw AS (
  SELECT Sale_date, net_revenue, hubspot_ID, formatted_date, dupe_flag,
    UPPER(webinar_type) AS wt,
    DATE(event_start_date_time, 'Asia/Kolkata') AS web_dt,
    ROW_NUMBER() OVER (PARTITION BY hubspot_ID, DATE(event_start_date_time, 'Asia/Kolkata') ORDER BY formatted_date ASC) AS rnk
  FROM `ik-marketing-data.India_Leads.US_Domain_combined_view`
  WHERE dupe_logic = 1
    AND COALESCE(Alumni_Stats, 'New_Lead') = 'New_Lead'
)
SELECT FORMAT_DATE('%Y-%m', s.Sale_date) AS sale_month,
       COUNT(*) AS sales_booked,
       SUM(s.net_revenue * COALESCE(fx.rate, 84.0)) AS revenue_booked
FROM sales_raw s
INNER JOIN events_tracked e ON s.wt = e.wt AND s.web_dt = e.web_dt
LEFT JOIN `{BQ_APP_PROJECT}.events.fx_rates_monthly` fx
  ON fx.month = FORMAT_DATE('%Y-%m', s.Sale_date)
  AND fx.base = 'USD' AND fx.quote = 'INR'
WHERE s.rnk = 1 AND s.dupe_flag = 0
  AND s.Sale_date IS NOT NULL AND s.Sale_date >= DATE '2026-01-01'
  AND COALESCE(s.net_revenue, 0) > 0
GROUP BY sale_month
"""
            else:  # US
                booked_query = f"""
WITH events_tracked AS (
  SELECT UPPER(webinar_type) AS wt, DATE(live_at, 'America/Los_Angeles') AS web_dt
  FROM `{BQ_APP_PROJECT}.events.event`
  WHERE LOWER(country) IN ('us','usa') AND live_at >= TIMESTAMP '2026-01-01 00:00:00'
    AND COALESCE(status, 'upcoming') != 'archived'
),
sales_raw AS (
  SELECT Sale_date, net_revenue, leads_hubspot_id, formatted_date, dupe_flag,
    UPPER(webinar_type) AS wt,
    DATE(event_start_date_time, 'America/Los_Angeles') AS web_dt,
    ROW_NUMBER() OVER (PARTITION BY leads_hubspot_id, DATE(event_start_date_time, 'America/Los_Angeles') ORDER BY formatted_date ASC) AS rnk
  FROM `ik-marketing-data.Marketing_data_new_logic.Bq_data_Alumni`
  WHERE dupe_logic = 1
    AND COALESCE(Alumni_Stats, 'New_Lead') = 'New_Lead'
)
SELECT FORMAT_DATE('%Y-%m', s.Sale_date) AS sale_month,
       COUNT(*) AS sales_booked,
       SUM(s.net_revenue) AS revenue_booked
FROM sales_raw s
INNER JOIN events_tracked e ON s.wt = e.wt AND s.web_dt = e.web_dt
WHERE s.rnk = 1 AND s.dupe_flag = 0
  AND s.Sale_date IS NOT NULL AND s.Sale_date >= DATE '2026-01-01'
  AND COALESCE(s.net_revenue, 0) > 0
GROUP BY sale_month
"""
            for r in src_client.query(booked_query).result():
                ms = r['sale_month']
                if ms in by_month:
                    by_month[ms]['revenue_booked'] = float(r['revenue_booked'] or 0.0)
                    by_month[ms]['sales_booked']   = int(r['sales_booked'] or 0)
        except Exception as e:
            print(f'[months/booked_revenue] {e}')

        months_out = sorted(by_month.values(), key=lambda m: m['month_slug'], reverse=True)
        self._json_response(200, {
            'country': country,
            'series':  series,
            'months':  months_out,
            'events':  events,  # drill-down data (per-event-within-month)
        }, extra_headers={'Cache-Control': 'private, max-age=60'})

    def _list_sales(self, parsed_url):
        """List individual sales attributable to a tracked masterclass event.

        US (USD): reads `Bq_data_Alumni`, dedupes by (leads_hubspot_id, PT date).
        India (INR): reads `US_Domain_combined_view`, dedupes by (hubspot_ID, IST
        date), filters out students + work_ex 0-2/3-4 (matching the snapshot's
        target-audience scope), and converts USD net_revenue to INR using each
        sale's month avg rate from `events.fx_rates_monthly` (fallback 84).

        Both branches INNER JOIN `events.event` on (webinar_type, local-TZ live_at)
        — only sales matching a tracked masterclass are returned. Sorted by
        Sale_date DESC; frontend groups by sale-month."""
        from google.cloud import bigquery
        qs       = urllib.parse.parse_qs(parsed_url.query)
        country  = qs.get('country', ['US'])[0]
        since    = qs.get('since', ['2026-01-01'])[0]
        c        = country.lower()
        if c in ('us', 'usa'):
            country_norm = 'US'
            currency     = 'USD'
            tz           = 'America/Los_Angeles'
            country_pred = "LOWER(e.country) IN ('us', 'usa')"
            revenue_expr = "s.net_revenue"
            extra_ctes   = ""
            extra_join   = ""
            extra_filter = ""
            source_table = "`ik-marketing-data.Marketing_data_new_logic.Bq_data_Alumni`"
            hubspot_col  = "leads_hubspot_id"
        elif c in ('india', 'in'):
            country_norm = 'India'
            currency     = 'INR'
            tz           = 'Asia/Kolkata'
            country_pred = "LOWER(e.country) = 'india'"
            # USD→INR at sale-month avg rate; fallback 84 (matches snapshot code)
            revenue_expr = "s.net_revenue * COALESCE(fx.rate, 84.0)"
            extra_ctes   = ""
            extra_join   = (
                f"LEFT JOIN `{BQ_APP_PROJECT}.events.fx_rates_monthly` fx\n"
                "  ON fx.month = FORMAT_DATE('%Y-%m', s.Sale_date)\n"
                "  AND fx.base = 'USD' AND fx.quote = 'INR'"
            )
            # No work_ex exclusion for sales — a paying customer of any experience
            # level should be counted, even if they were outside the target-audience
            # filter used elsewhere (calls / IQLs / etc).
            extra_filter = ""
            source_table = "`ik-marketing-data.India_Leads.US_Domain_combined_view`"
            hubspot_col  = "hubspot_ID"
        else:
            self._json_response(400, {'error': 'country must be US or India'})
            return

        query = f"""
WITH sales_raw AS (
  SELECT
    Sale_date, net_revenue, Channel, webinar_type,
    DATE(event_start_date_time, "{tz}") AS web_scheduled_date,
    {hubspot_col} AS hubspot_id, formatted_date, dupe_flag,
    ROW_NUMBER() OVER (
      PARTITION BY {hubspot_col}, DATE(event_start_date_time, "{tz}")
      ORDER BY formatted_date ASC
    ) AS rnk
  FROM {source_table}
  WHERE dupe_logic = 1
    AND COALESCE(Alumni_Stats, 'New_Lead') = 'New_Lead'   -- exclude alumni-repeat sales
    {extra_filter}
),
sales AS (
  SELECT Sale_date, net_revenue, Channel, webinar_type, web_scheduled_date
  FROM sales_raw
  WHERE rnk = 1
    AND dupe_flag = 0
    AND Sale_date IS NOT NULL
    AND Sale_date >= DATE(@since)
    AND COALESCE(net_revenue, 0) > 0
)
SELECT
  s.Sale_date,
  {revenue_expr} AS revenue,
  s.Channel,
  s.webinar_type,
  s.web_scheduled_date,
  e.event_id,
  e.title       AS event_title,
  e.instructor_name
FROM sales s
JOIN `{BQ_APP_PROJECT}.events.event` e
  ON e.webinar_type = s.webinar_type
  AND DATE(e.live_at, "{tz}") = s.web_scheduled_date
  AND {country_pred}
{extra_join}
WHERE COALESCE(e.status, 'upcoming') != 'archived'
ORDER BY s.Sale_date DESC, revenue DESC"""
        params = [bigquery.ScalarQueryParameter('since', 'DATE', since)]
        try:
            client = bigquery.Client(project=BQ_APP_PROJECT)
            rows   = list(client.query(query, job_config=bigquery.QueryJobConfig(query_parameters=params)).result())
        except Exception as e:
            self._json_response(500, {'error': str(e)})
            return

        sales = []
        for r in rows:
            d = _row_to_dict(r)
            sales.append({
                'sale_date':          d.get('Sale_date'),
                'revenue':            float(d['revenue']) if d.get('revenue') is not None else 0.0,
                'channel':            d.get('Channel') or 'Unknown',
                'webinar_type':       d.get('webinar_type'),
                'web_scheduled_date': d.get('web_scheduled_date'),
                'event_id':           d.get('event_id'),
                'event_title':        d.get('event_title'),
                'instructor_name':    d.get('instructor_name'),
            })
        self._json_response(200, {
            'country':  country_norm,
            'currency': currency,
            'since':    since,
            'sales':    sales,
        }, extra_headers={'Cache-Control': 'private, max-age=60'})

    def _get_event(self, event_id):
        from google.cloud import bigquery
        if not event_id or '/' in event_id:
            self._json_response(400, {'error': 'Invalid event_id'})
            return
        try:
            client = bigquery.Client(project=BQ_APP_PROJECT)
            param  = [bigquery.ScalarQueryParameter('eid', 'STRING', event_id)]

            def run(sql, params):
                return list(client.query(sql, job_config=bigquery.QueryJobConfig(query_parameters=params)).result())

            with ThreadPoolExecutor(max_workers=3) as ex:
                ev_fut = ex.submit(run,
                    f"SELECT * FROM `{BQ_APP_PROJECT}.events.event` WHERE event_id = @eid", param)
                snap_fut = ex.submit(run,
                    f"""SELECT * FROM `{BQ_APP_PROJECT}.history.event_snapshot`
                        WHERE event_id = @eid
                        ORDER BY snapshot_at DESC LIMIT 1""", param)
                # Align daily rows with the same snapshot run as the cumulative
                # snapshot above. Without this, a refresh that writes fewer dates
                # than a previous run leaves stale per-date rows from older runs,
                # so daily-sum > cumulative (bento KPI). Filtering by the latest
                # snapshot_at guarantees both views come from one point in time.
                daily_fut = ex.submit(run,
                    f"""SELECT * FROM `{BQ_APP_PROJECT}.history.event_daily`
                        WHERE event_id = @eid
                          AND snapshot_at = (
                            SELECT MAX(snapshot_at)
                            FROM `{BQ_APP_PROJECT}.history.event_snapshot`
                            WHERE event_id = @eid
                          )
                        ORDER BY registration_date""", param)

                ev_rows = ev_fut.result()
                if not ev_rows:
                    self._json_response(404, {'error': 'Event not found'})
                    return
                event      = _row_to_dict(ev_rows[0])
                snap_rows  = snap_fut.result()
                daily_rows = daily_fut.result()

            snapshot = _row_to_dict(snap_rows[0]) if snap_rows else None
            daily    = [_row_to_dict(r) for r in daily_rows]

            # Real per-channel attendance — computed at read-time only for aired
            # events. Joins Zoom attendees to lead channel/utm. Failure is
            # non-fatal; UI falls back to overall-rate approximation.
            channel_attendance = None
            try:
                if event.get('live_at') and event.get('webinar_type'):
                    live_at = event['live_at']
                    if isinstance(live_at, str):
                        live_at_dt = datetime.datetime.fromisoformat(live_at.replace('Z', '+00:00'))
                    else:
                        live_at_dt = live_at
                    now_utc = datetime.datetime.now(datetime.timezone.utc)
                    aware   = live_at_dt if live_at_dt.tzinfo else live_at_dt.replace(tzinfo=datetime.timezone.utc)
                    if aware <= now_utc:
                        channel_attendance = _query_channel_attendance(event)
            except Exception as e:
                print(f'[channel_attendance] {event_id}: {e}')

            self._json_response(200, {
                'event': event,
                'snapshot': snapshot,
                'daily': daily,
                'channel_attendance': channel_attendance,
            }, extra_headers={'Cache-Control': 'private, max-age=60'})
        except Exception as e:
            self._json_response(500, {'error': str(e)})

    def _create_event(self):
        email  = _session_email(self) or 'unknown'
        length = int(self.headers.get('Content-Length', 0))
        body   = json.loads(self.rfile.read(length)) if length else {}

        try:
            from google.cloud import bigquery
            client   = bigquery.Client(project=BQ_APP_PROJECT)
            event_id = _generate_event_id(body)
            now_iso  = datetime.datetime.utcnow().replace(microsecond=0).isoformat() + 'Z'

            row = {
                'event_id':         event_id,
                'title':            body.get('title') or '',
                'topic':            body.get('topic') or None,
                'event_type':       body.get('eventType') or None,
                'country':          body.get('country') or None,
                'webinar_type':     body.get('webinarType') or None,
                'live_at':          body.get('liveAt') or None,
                'day2_live_at':     body.get('day2LiveAt') or None,
                'go_live_date':     body.get('goLiveDate') or None,
                'landing_url':      body.get('landingUrl') or None,
                'zoom_url':         body.get('zoomUrl') or None,
                'slides_url':       body.get('slidesUrl') or None,
                'yt_url':           body.get('ytUrl') or None,
                'instructor_name':  body.get('instructorName') or None,
                'instructor_role':  body.get('instructorRole') or None,
                'summary':          body.get('summary') or None,
                'design_notes':     body.get('designNotes') or None,
                'goal_regs':        body.get('goalRegs'),
                'status':           'upcoming',
                'jira_design_key':  body.get('jiraDesignKey') or None,
                'jira_landing_key': body.get('jiraLandingKey') or None,
                'created_at':       now_iso,
                'created_by':       email,
                'updated_at':       now_iso,
            }

            table  = f'{BQ_APP_PROJECT}.events.event'
            errors = client.insert_rows_json(table, [row])
            if errors:
                self._json_response(500, {'error': 'BigQuery insert failed', 'details': errors})
                return

            self._json_response(200, {'ok': True, 'event_id': event_id})

        except ImportError:
            self._json_response(500, {'error': 'google-cloud-bigquery not installed'})
        except Exception as e:
            self._json_response(500, {'error': str(e)})

    def _save_snapshot_config(self):
        length = int(self.headers.get('Content-Length', 0))
        body = json.loads(self.rfile.read(length)) if length else {}
        try:
            save_snapshot_config(body)
            self._json_response(200, {'ok': True})
        except Exception as e:
            self._json_response(500, {'error': str(e)})

    def _poll_qna_export(self, parsed_url):
        """Pull poll + Q&A rows for an event from the Zoom views and return an
        XLSX file with two sheets (Polls, Q&A) as a direct download. No Google
        Drive / Sheets API needed — generates the file in memory with openpyxl."""
        qs = urllib.parse.parse_qs(parsed_url.query)
        event_id = qs.get('event_id', [None])[0] or qs.get('id', [None])[0]
        if not event_id:
            self._json_response(400, {'error': 'Missing event_id'})
            return
        try:
            from google.cloud import bigquery
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            import io, re

            client = bigquery.Client(project=BQ_APP_PROJECT)
            ev_rows = list(client.query(
                f"SELECT title, live_at, country FROM `{BQ_APP_PROJECT}.events.event` WHERE event_id = @eid",
                job_config=bigquery.QueryJobConfig(query_parameters=[bigquery.ScalarQueryParameter('eid', 'STRING', event_id)])
            ).result())
            if not ev_rows:
                self._json_response(404, {'error': f'Event not found: {event_id}'})
                return
            ev = ev_rows[0]
            title = ev['title'] or ''
            live_at = ev['live_at']
            country = (ev['country'] or '').strip()
            is_us = country.lower() in ('us', 'usa', 'united states')
            tz_name = 'America/Los_Angeles' if is_us else 'Asia/Kolkata'
            date_col = 'webinar_date_pst' if is_us else 'webinar_date_ist'

            # live_at is a UTC timestamp; convert to event's local tz to get the right calendar date.
            if live_at is None:
                self._json_response(400, {'error': 'Event has no live_at date'})
                return
            try:
                import zoneinfo
                local_tz = zoneinfo.ZoneInfo(tz_name)
            except Exception:
                local_tz = datetime.timezone(datetime.timedelta(hours=5, minutes=30) if not is_us else datetime.timedelta(hours=-7))
            event_date = live_at.astimezone(local_tz).date()
            # Search ±1 day to absorb Zoom's date-capture timing quirks
            d_lo = (event_date - datetime.timedelta(days=1)).isoformat()
            d_hi = (event_date + datetime.timedelta(days=1)).isoformat()

            # Fuzzy keyword from the event title — use the leading distinctive phrase
            # (drop trailing dash/em-dash suffixes like " — Master AI Engineering").
            import re
            base = re.split(r'\s*[—–-]\s+', title, maxsplit=1)[0].strip()
            keyword = base[:40] if base else title[:30]

            src = bigquery.Client(project='ik-marketing-data')

            poll_q = f"""SELECT * FROM `ik-marketing-data.Webinar_analytics.zoom_webinar_polls_view`
              WHERE DATE(webinar_date, '{tz_name}') BETWEEN @lo AND @hi
                AND LOWER(IFNULL(webinar_topic, '')) LIKE CONCAT('%', LOWER(@k), '%')"""
            poll_job = src.query(poll_q, job_config=bigquery.QueryJobConfig(query_parameters=[
                bigquery.ScalarQueryParameter('lo', 'DATE', d_lo),
                bigquery.ScalarQueryParameter('hi', 'DATE', d_hi),
                bigquery.ScalarQueryParameter('k', 'STRING', keyword),
            ]))
            poll_rows = list(poll_job.result())
            poll_fields = [f.name for f in poll_job.schema] if poll_job.schema else []

            qna_q = f"""SELECT * FROM `ik-marketing-data.Webinar_analytics.zoom_webinar_qa_json_view`
              WHERE DATE(webinar_date, '{tz_name}') BETWEEN @lo AND @hi
                AND LOWER(IFNULL(webinar_type, '')) LIKE CONCAT('%', LOWER(@k), '%')"""
            qna_job = src.query(qna_q, job_config=bigquery.QueryJobConfig(query_parameters=[
                bigquery.ScalarQueryParameter('lo', 'DATE', d_lo),
                bigquery.ScalarQueryParameter('hi', 'DATE', d_hi),
                bigquery.ScalarQueryParameter('k', 'STRING', keyword),
            ]))
            qna_rows = list(qna_job.result())
            qna_fields = [f.name for f in qna_job.schema] if qna_job.schema else []

            wb = Workbook()
            header_font = Font(bold=True, color='18181B')
            header_fill = PatternFill(start_color='F0F0F2', end_color='F0F0F2', fill_type='solid')
            header_align = Alignment(horizontal='left', vertical='center')

            def populate(ws, fields, rows, label):
                if not fields:
                    ws.append(['No schema available'])
                    return
                ws.append(fields)
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                ws.freeze_panes = 'A2'
                for r in rows:
                    ws.append([_cell_to_str(r.get(f)) for f in fields])
                if not rows:
                    ws.append([f'(no {label} matched "{title}" on {event_date.isoformat()})'])
                # Reasonable column widths
                for i, fname in enumerate(fields, 1):
                    ws.column_dimensions[get_column_letter(i)].width = min(40, max(12, len(fname) + 2))

            # First sheet
            ws_p = wb.active
            ws_p.title = 'Polls'
            populate(ws_p, poll_fields, poll_rows, 'polls')
            # Second sheet
            ws_q = wb.create_sheet('Q&A')
            populate(ws_q, qna_fields, qna_rows, 'Q&A')

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            xlsx_bytes = buf.getvalue()

            safe_title = re.sub(r'[^A-Za-z0-9._-]+', '_', title)[:60]
            filename = f"polls-qna_{safe_title}_{event_date.isoformat()}.xlsx"

            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
            self.send_header('X-Poll-Rows', str(len(poll_rows)))
            self.send_header('X-Qna-Rows', str(len(qna_rows)))
            self.send_header('Content-Length', str(len(xlsx_bytes)))
            self.end_headers()
            self.wfile.write(xlsx_bytes)
        except Exception as e:
            import traceback, sys
            tb = traceback.format_exc()
            sys.stderr.write(f'[poll-qna-export] error: {e}\n{tb}\n')
            sys.stderr.flush()
            self._json_response(500, {'error': str(e), 'traceback': tb.splitlines()[-6:]})

    def _run_snapshot(self):
        try:
            qs = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
            days_back = int(qs.get('days_back', ['14'])[0])
            _do_snapshot(days_back=days_back)
            self._json_response(200, {'ok': True, 'days_back': days_back})
        except Exception as e:
            self._json_response(500, {'error': str(e)})

    def _refresh_event(self, event_id):
        """Per-event snapshot. Runs the same _build_event_snapshot path the cron
        uses, but for a single event_id picked by the caller. Skips the cron's
        active-window filter so the user can refresh older events too.
        Writes new rows to history.event_daily + history.event_snapshot."""
        from google.cloud import bigquery
        if not event_id or '/' in event_id:
            self._json_response(400, {'error': 'Invalid event_id'})
            return
        try:
            client = bigquery.Client(project=BQ_APP_PROJECT)
            rows = list(client.query(
                f"SELECT event_id, country, webinar_type, live_at, day2_live_at "
                f"FROM `{BQ_APP_PROJECT}.events.event` WHERE event_id = @eid",
                job_config=bigquery.QueryJobConfig(query_parameters=[
                    bigquery.ScalarQueryParameter('eid', 'STRING', event_id)
                ])
            ).result())
            if not rows:
                self._json_response(404, {'error': 'Event not found'})
                return
            ev = rows[0]
            now = datetime.datetime.now(datetime.timezone.utc).replace(microsecond=0)
            now_iso = now.isoformat().replace('+00:00', 'Z')
            result = _build_event_snapshot(client, ev, now, now_iso)
            if result is None:
                self._json_response(400, {'error': f'Country "{ev.country}" snapshot not supported. India and US only.'})
                return
            daily_rows, snap_row = result
            if daily_rows:
                errs = client.insert_rows_json(f'{BQ_APP_PROJECT}.history.event_daily', daily_rows)
                if errs:
                    self._json_response(500, {'error': 'event_daily insert failed', 'details': errs})
                    return
            errs = client.insert_rows_json(f'{BQ_APP_PROJECT}.history.event_snapshot', [snap_row])
            if errs:
                self._json_response(500, {'error': 'event_snapshot insert failed', 'details': errs})
                return
            self._json_response(200, {
                'ok': True,
                'event_id': event_id,
                'daily_rows': len(daily_rows),
                'snapshot_at': snap_row.get('snapshot_at'),
            })
        except Exception as e:
            import traceback, sys
            tb = traceback.format_exc()
            sys.stderr.write(f'[refresh-event] {event_id} error: {e}\n{tb}\n')
            sys.stderr.flush()
            self._json_response(500, {'error': str(e), 'traceback': tb.splitlines()[-6:]})

    def _auth_verify(self):
        length = int(self.headers.get('Content-Length', 0))
        body   = json.loads(self.rfile.read(length)) if length else {}
        credential = body.get('credential', '')
        if not GOOGLE_CLIENT_ID:
            self._json_response(500, {'error': 'GOOGLE_CLIENT_ID not configured on server.'})
            return
        try:
            from google.oauth2 import id_token
            from google.auth.transport import requests as grequests
            idinfo = id_token.verify_oauth2_token(credential, grequests.Request(), GOOGLE_CLIENT_ID)
            email  = idinfo.get('email', '')
            hd     = idinfo.get('hd', '')
            if not (hd == ALLOWED_DOMAIN or email.endswith('@' + ALLOWED_DOMAIN)):
                self._json_response(403, {'error': 'Only @interviewkickstart.com accounts are allowed.'})
                return
            token = _make_session(email)
            self.send_response(200)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Set-Cookie',
                'ik_session=' + token + '; Path=/; HttpOnly; SameSite=Strict; Max-Age=' + str(SESSION_TTL))
            self.end_headers()
            self.wfile.write(json.dumps({'ok': True, 'email': email}).encode())
        except Exception as e:
            self._json_response(403, {'error': str(e)})

    def _auth_logout(self):
        self.send_response(200)
        self._cors()
        self.send_header('Content-Type', 'application/json')
        self.send_header('Set-Cookie', 'ik_session=; Path=/; HttpOnly; SameSite=Strict; Max-Age=0')
        self.end_headers()
        self.wfile.write(json.dumps({'ok': True}).encode())

    def log_message(self, fmt, *args):
        print(f"  {args[0]} {args[1]}")

SNAPSHOT_CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'snapshot_config.json')

def load_snapshot_config():
    if GCS_BUCKET:
        try:
            from google.cloud import storage
            blob = storage.Client().bucket(GCS_BUCKET).blob(GCS_CONFIG_KEY)
            return json.loads(blob.download_as_text())
        except Exception:
            return {}
    try:
        with open(SNAPSHOT_CONFIG_FILE) as f:
            return json.load(f)
    except Exception:
        return {}

def save_snapshot_config(body):
    if GCS_BUCKET:
        from google.cloud import storage
        blob = storage.Client().bucket(GCS_BUCKET).blob(GCS_CONFIG_KEY)
        blob.upload_from_string(json.dumps(body, indent=2), content_type='application/json')
    else:
        with open(SNAPSHOT_CONFIG_FILE, 'w') as f:
            json.dump(body, f, indent=2)

def build_snapshot_message(rows, schema):
    col_map = {name: i for i, name in enumerate(schema)}
    date_idx  = col_map.get('registration_date', 0)
    chan_idx   = col_map.get('channel', 1)
    count_idx  = col_map.get('total_leads', 3)
    spend_idx  = col_map.get('total_spend', 4)

    buckets = {'facebook': {}, 'crm': {}, 'others': {}, 'overall': {}}
    spends  = {}
    seen    = set()

    def is_fb(ch):  return bool(__import__('re').search(r'facebook|fb', ch or '', flags=2))
    def is_crm(ch): return bool(__import__('re').search(r'email|whatsapp|whats.?app', ch or '', flags=2))

    for row in rows:
        d     = str(row[date_idx] or '')[:10]
        ch    = str(row[chan_idx] or '')
        count = float(row[count_idx] or 0)
        spend = float(row[spend_idx] or 0)
        if spend and d not in seen:
            seen.add(d)
            spends[d] = spends.get(d, 0) + spend
        buckets['overall'][d] = buckets['overall'].get(d, 0) + count
        if is_fb(ch):   buckets['facebook'][d] = buckets['facebook'].get(d, 0) + count
        elif is_crm(ch):buckets['crm'][d]      = buckets['crm'].get(d, 0) + count
        else:            buckets['others'][d]   = buckets['others'].get(d, 0) + count

    all_dates = sorted(buckets['overall'].keys())
    lines = ['*Registration Summary*', '```']
    hdr = f"{'Date':<12}{'Spends (₹)':>14}{'Meta':>7}{'CPIQL':>8}{'CRM':>6}{'Others':>8}{'Overall':>9}"
    lines.append(hdr)
    lines.append('-' * 64)
    t = {'spends': 0, 'meta': 0, 'crm': 0, 'others': 0, 'overall': 0}
    for d in all_dates:
        sp  = spends.get(d, 0)
        meta= buckets['facebook'].get(d, 0)
        crm = buckets['crm'].get(d, 0)
        oth = buckets['others'].get(d, 0)
        all_= buckets['overall'].get(d, 0)
        cpiql = f'{sp/meta:.2f}' if meta > 0 else '—'
        sp_fmt = f'{round(sp):,}' if sp else '—'
        lines.append(f"{d:<12}{sp_fmt:>14}{int(meta):>7}{cpiql:>8}{int(crm):>6}{int(oth):>8}{int(all_):>9}")
        t['spends'] += sp; t['meta'] += meta; t['crm'] += crm; t['others'] += oth; t['overall'] += all_
    lines.append('-' * 64)
    tc = f"{t['spends']/t['meta']:.2f}" if t['meta'] > 0 else '—'
    total_sp_fmt = f"{round(t['spends']):,}"
    lines.append(f"{'Total':<12}{total_sp_fmt:>14}{int(t['meta']):>7}{tc:>8}{int(t['crm']):>6}{int(t['others']):>8}{int(t['overall']):>9}")
    lines.append('```')
    return '\n'.join(lines)

def send_slack_dm(token, user_id, text):
    def post(url, body):
        req = urllib.request.Request(url, data=json.dumps(body).encode(),
              headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}, method='POST')
        with urllib.request.urlopen(req, context=SSL_CTX) as r:
            return json.loads(r.read())
    ch = post('https://slack.com/api/conversations.open', {'users': user_id})
    if not ch.get('ok'): raise Exception(ch.get('error'))
    post('https://slack.com/api/chat.postMessage', {'channel': ch['channel']['id'], 'text': text})

def _do_snapshot(days_back=14):
    """Per-event BQ snapshot — writes to history.event_daily + history.event_snapshot.
    days_back widens the active-event window for backfilling older events.

    The legacy Slack DM path (_snapshot_to_slack) is currently DISABLED — its
    spend-join logic uses naive LIKE matching while _snapshot_events_to_bq uses
    REGEXP-normalised matching, producing slightly different numbers in the
    daily Slack DM vs the event page. Re-enable below once the dedup logic is
    ported from server.py:1024-1034. Manual "Send to Slack" on the legacy Leads
    tab still works (separate code path)."""
    errors = []
    # try:
    #     _snapshot_to_slack()
    # except Exception as e:
    #     errors.append(f'slack: {e}')
    #     print(f'  [snapshot/slack] error: {e}')
    try:
        _snapshot_events_to_bq(days_back=days_back)
    except Exception as e:
        errors.append(f'bq: {e}')
        print(f'  [snapshot/bq] error: {e}')
    if errors:
        raise Exception('; '.join(errors))

def _snapshot_to_slack():
    """Legacy daily Slack DM driven by snapshot_config (will be replaced in slice 3+)."""
    cfg = load_snapshot_config()
    if not cfg.get('slackToken') or not cfg.get('slackSnapshotId') or not cfg.get('webinarDates'):
        raise Exception('snapshot config missing')
    from google.cloud import bigquery
    date_literal = ', '.join(f"'{d}'" for d in cfg['webinarDates'])
    webinar_type = cfg.get('webinarType', '')
    query = f"""
WITH leads AS (
  SELECT DATE(formatted_date) AS registration_date, channel, utm_campaign, COUNT(*) AS total_leads
  FROM (
    SELECT formatted_date, channel, utm_campaign,
      DATE(event_start_date_time, "Asia/Kolkata") AS web_scheduled_date,
      hubspot_ID, dupe_flag, gql_flag, work_ex, webinar_type, dupe_logic,
      ROW_NUMBER() OVER (PARTITION BY hubspot_ID, DATE(event_start_date_time, "Asia/Kolkata") ORDER BY formatted_date ASC) AS rnk
    FROM `ik-marketing-data.India_Leads.US_Domain_combined_view`
    WHERE dupe_logic = 1
  )
  WHERE rnk = 1 AND web_scheduled_date IN ({date_literal}) AND webinar_type = '{webinar_type}'
    AND dupe_flag = 0 AND gql_flag = 0
    AND LOWER(work_ex) NOT LIKE '%student%' AND work_ex NOT IN ('0-2', '3-4')
  GROUP BY registration_date, channel, utm_campaign
),
spends AS (
  SELECT DATE(campaign_date) AS spend_date, campaign_name, SUM(cost) AS total_spend
  FROM `ik-marketing-data.India_Leads.Combined_India_Spend`
  -- Same exclusions as US (2026-06-22), kept for parity if this legacy
  -- _snapshot_to_slack path is ever re-enabled.
  WHERE (LOWER(campaign_name) LIKE '%l10x%' OR LOWER(campaign_name) LIKE '%masterclass%')
    AND (LOWER(campaign_name) LIKE '%meta%' OR LOWER(campaign_name) LIKE '%facebook%' OR LOWER(campaign_name) LIKE '%l10x%')
    AND LOWER(campaign_name) NOT LIKE '%recorded_masterclass%'
    AND LOWER(campaign_name) NOT LIKE '%holiday_offer%'
  GROUP BY spend_date, campaign_name
)
SELECT l.registration_date, l.channel, l.utm_campaign, l.total_leads, SUM(s.total_spend) AS total_spend
FROM leads l
LEFT JOIN spends s ON l.registration_date = s.spend_date
  AND LOWER(l.utm_campaign) LIKE CONCAT('%', LOWER(s.campaign_name), '%')
GROUP BY l.registration_date, l.channel, l.utm_campaign, l.total_leads
ORDER BY l.registration_date, l.channel, l.utm_campaign"""

    client = bigquery.Client(project='ik-marketing-data')
    result = client.query(query).result()
    schema = [f.name for f in result.schema]
    rows   = [list(row.values()) for row in result]
    msg = build_snapshot_message(rows, schema)
    send_slack_dm(cfg['slackToken'], cfg['slackSnapshotId'], msg)
    print(f'  [snapshot/slack] sent to {cfg["slackSnapshotId"]}')

# ─── Per-event BQ snapshots ──────────────────────────────────────────────────

# Channel classifiers — match the existing Slack snapshot logic.
_RE_META = re.compile(r'facebook|fb', re.I)
_RE_CRM  = re.compile(r'email|whatsapp|whats.?app', re.I)

# Per-event spend corrections — added when source pipeline drops days.
# Survives nightly cron overwrites: applied inside _cumulative_snapshot_row.
# meta_spend is in the event's local currency (INR for India, USD for US).
SPEND_CORRECTIONS = {
    # May 3, 2026 spend was dropped from Combined_India_Spend for the 5
    # Pilot_Meta_L10X_India_01_AI_Launchpad_9_10_May* campaigns.
    # User-verified (2026-06-23) from Meta Ads Manager:
    #   May1 ₹14,842 + May2 ₹9,783 + May3 ₹12,994 ≈ ₹37,619.
    'ai-launchpad-master-ai-may09': 37619.0,
}

def _snapshot_events_to_bq(days_back=14):
    """For each active event, query lead funnel and write rows to
    history.event_daily (per registration_date) + history.event_snapshot (cumulative).
    days_back controls the past-event lookback (default 14d); pass a larger value
    via /run-snapshot?days_back=N to backfill older events."""
    from google.cloud import bigquery
    client  = bigquery.Client(project=BQ_APP_PROJECT)
    now     = datetime.datetime.now(datetime.timezone.utc).replace(microsecond=0)
    now_iso = now.isoformat().replace('+00:00', 'Z')

    events = list(client.query(f"""
        SELECT event_id, country, webinar_type, live_at, day2_live_at
        FROM `{BQ_APP_PROJECT}.events.event`
        WHERE COALESCE(status, 'upcoming') != 'archived'
          AND webinar_type IS NOT NULL
          AND live_at IS NOT NULL
          AND live_at BETWEEN TIMESTAMP_SUB(CURRENT_TIMESTAMP(), INTERVAL {int(days_back)} DAY)
                          AND TIMESTAMP_ADD(CURRENT_TIMESTAMP(), INTERVAL 60 DAY)
    """).result())

    print(f'  [snapshot/bq] {len(events)} active event(s)')

    # Insert per-event so partial progress survives timeouts/errors.
    for ev in events:
        try:
            result = _build_event_snapshot(client, ev, now, now_iso)
            if result is None:
                print(f'  [snapshot/bq] {ev.event_id}: country={ev.country} not supported yet, skipping')
                continue
            d, s = result
            if d:
                errs = client.insert_rows_json(f'{BQ_APP_PROJECT}.history.event_daily', d)
                if errs: print(f'  [snapshot/bq] {ev.event_id} event_daily errors: {errs}')
            errs = client.insert_rows_json(f'{BQ_APP_PROJECT}.history.event_snapshot', [s])
            if errs: print(f'  [snapshot/bq] {ev.event_id} event_snapshot errors: {errs}')
            print(f'  [snapshot/bq] {ev.event_id}: {len(d)} daily row(s), total_regs={s["total_regs"]}')
        except Exception as e:
            print(f'  [snapshot/bq] {ev.event_id}: error: {e}')

def _build_event_snapshot(client, ev, now, now_iso):
    """Dispatch to country-specific builder. Each returns (daily_rows, snapshot_row)
    or None if country unsupported. Each per-event query is wrapped in try/except
    so a single failure doesn't blank out the whole row."""
    c = (ev.country or '').lower()
    if c == 'india':
        return _build_event_snapshot_india(client, ev, now, now_iso)
    if c in ('us', 'usa'):
        return _build_event_snapshot_us(client, ev, now, now_iso)
    return None

def _build_event_snapshot_india(client, ev, now, now_iso):
    from google.cloud import bigquery

    live_dates = []
    for d in [ev.live_at, ev.day2_live_at]:
        if d is None: continue
        live_dates.append(d.date() if isinstance(d, datetime.datetime) else d)
    if not live_dates:
        return [], _cumulative_snapshot_row(ev.event_id, {}, ev.live_at, now, now_iso)

    date_literal = ', '.join(f"DATE '{d.isoformat()}'" for d in live_dates)
    src_client   = bigquery.Client(project='ik-marketing-data')
    wt_param     = [bigquery.ScalarQueryParameter('wt', 'STRING', ev.webinar_type)]

    # ─── 1. Leads + spends (per-day buckets) ──────────────────────────────────
    # Daily Meta spend is computed once per (date, campaign_name) via DISTINCT —
    # without this, a campaign that matches multiple utm_campaigns gets counted
    # multiple times. Previous "dedup-by-date in Python" was the opposite bug:
    # only one campaign per day was kept and the rest were silently dropped.
    query = f"""
WITH leads AS (
  SELECT DATE(formatted_date) AS registration_date, channel, utm_campaign, COUNT(*) AS total_leads
  FROM (
    SELECT formatted_date, channel, utm_campaign,
      DATE(event_start_date_time, "Asia/Kolkata") AS web_scheduled_date,
      hubspot_ID, dupe_flag, gql_flag, work_ex, webinar_type, dupe_logic,
      ROW_NUMBER() OVER (PARTITION BY hubspot_ID, DATE(event_start_date_time, "Asia/Kolkata") ORDER BY formatted_date ASC) AS rnk
    FROM `ik-marketing-data.India_Leads.US_Domain_combined_view`
    WHERE dupe_logic = 1
  )
  WHERE rnk = 1 AND web_scheduled_date IN ({date_literal}) AND UPPER(webinar_type) = UPPER(@wt)
    AND dupe_flag = 0 AND gql_flag = 0
    AND LOWER(work_ex) NOT LIKE '%student%' AND work_ex NOT IN ('0-2', '3-4')
  GROUP BY registration_date, channel, utm_campaign
),
spends AS (
  -- Dedup exact-duplicate rows in Combined_India_Spend before summing.
  -- Meta ad-report rows are uniquely keyed by (campaign_date, campaign_name,
  -- ad_group_id, device); duplicates appear when the source ETL re-runs an
  -- ingest without truncating. Confirmed on 2026-06-30 which had 2× dupes.
  SELECT spend_date, campaign_name, SUM(cost) AS total_spend
  FROM (
    SELECT DATE(campaign_date) AS spend_date, campaign_name, ad_group_id, device, ANY_VALUE(cost) AS cost
    FROM `ik-marketing-data.India_Leads.Combined_India_Spend`
    WHERE (LOWER(campaign_name) LIKE '%l10x%' OR LOWER(campaign_name) LIKE '%masterclass%')
      AND (LOWER(campaign_name) LIKE '%meta%' OR LOWER(campaign_name) LIKE '%facebook%' OR LOWER(campaign_name) LIKE '%l10x%')
      AND LOWER(campaign_name) NOT LIKE '%recorded_masterclass%'
      AND LOWER(campaign_name) NOT LIKE '%holiday_offer%'
    GROUP BY spend_date, campaign_name, ad_group_id, device
  )
  GROUP BY spend_date, campaign_name
),
matched_campaigns AS (
  -- Normalize both sides to alphanumeric-only before LIKE: utm_campaign URL-encodes
  -- spaces as '+', while ad-manager campaign_name keeps actual spaces (and may
  -- contain en-dashes, hyphens, parens). Stripping non-alphanumerics on both sides
  -- prevents trivial encoding mismatches from dropping legitimate matches.
  SELECT DISTINCT l.registration_date, s.campaign_name, s.total_spend
  FROM leads l
  JOIN spends s ON l.registration_date = s.spend_date
    AND REGEXP_REPLACE(LOWER(l.utm_campaign), r'[^a-z0-9]', '')
        LIKE CONCAT('%', REGEXP_REPLACE(LOWER(s.campaign_name), r'[^a-z0-9]', ''), '%')
),
daily_spend AS (
  SELECT registration_date, SUM(total_spend) AS day_spend
  FROM matched_campaigns
  GROUP BY registration_date
)
SELECT l.registration_date, l.channel, l.total_leads, COALESCE(ds.day_spend, 0) AS total_spend
FROM leads l
LEFT JOIN daily_spend ds ON l.registration_date = ds.registration_date
ORDER BY l.registration_date, l.channel"""

    job = src_client.query(query, job_config=bigquery.QueryJobConfig(query_parameters=wt_param))
    rows = list(job.result())

    # Bucket per (date, channel-class). Spend dedup per date matches old Slack logic.
    daily = {}
    spend_seen = set()
    for r in rows:
        d = r['registration_date'].isoformat() if r['registration_date'] else None
        if not d: continue
        ch    = r['channel'] or ''
        count = float(r['total_leads'] or 0)
        spend = float(r['total_spend'] or 0)

        b = daily.setdefault(d, {
            'meta_regs': 0.0, 'meta_spend': 0.0,
            'crm_regs': 0.0,
            'other_regs': 0.0, 'other_spend': 0.0,
        })
        if spend and d not in spend_seen:
            spend_seen.add(d)
            b['meta_spend'] += spend  # source query only pulls meta-flavored spends
        if   _RE_META.search(ch): b['meta_regs']  += count
        elif _RE_CRM.search(ch):  b['crm_regs']   += count
        else:                     b['other_regs'] += count

    daily_rows = []
    for d, b in sorted(daily.items()):
        total_regs  = b['meta_regs'] + b['crm_regs'] + b['other_regs']
        total_spend = b['meta_spend'] + b['other_spend']
        cpiql       = (b['meta_spend'] / b['meta_regs']) if b['meta_regs'] > 0 else None
        daily_rows.append({
            'event_id':          ev.event_id,
            'registration_date': d,
            'snapshot_at':       now_iso,
            'meta_regs':         int(b['meta_regs']),
            'meta_spend':        b['meta_spend'],
            'crm_regs':          int(b['crm_regs']),
            'other_regs':        int(b['other_regs']),
            'other_spend':       b['other_spend'],
            'total_regs':        int(total_regs),
            'total_spend':       total_spend,
            'cpiql':             cpiql,
            'extras':            None,
        })

    # ─── 2-6. Sub-queries (cohort, attendance, calls, emails, sales) in parallel ──
    # Each is an independent BigQuery scan; running them concurrently drops
    # per-event snapshot time from sum(queries) to max(queries). _safe_query
    # swallows per-query failures so one slow source doesn't take down the row.
    cohort_default = {'role_sde': None, 'role_ml': None, 'role_fe': None,
                      'role_management': None, 'role_systems': None, 'role_null': None, 'role_other': None,
                      'we_0_2': None, 'we_3_5': None, 'we_6_10': None,
                      'we_10_15': None, 'we_15_20': None, 'we_20p': None,
                      'we_other': None, 'we_10p': None}
    with ThreadPoolExecutor(max_workers=5) as ex:
        f_cohort     = ex.submit(_safe_query, _query_cohort_india,     src_client, date_literal, wt_param, label=f'cohort/{ev.event_id}',     default=cohort_default)
        f_attendance = ex.submit(_safe_query, _query_attendance_india, src_client, date_literal, wt_param, label=f'attendance/{ev.event_id}', default={'attendees': None, 'attendance_pct': None})
        f_calls      = ex.submit(_safe_query, _query_calls_india,      src_client, date_literal, wt_param, label=f'calls/{ev.event_id}',      default={
            'calls_attempted': None, 'calls_connected': None, 'avg_talk_seconds': None,
            'call_total_leads': None,
            'call_pre_attempts': None, 'call_pre_connects': None, 'call_pre_talk_mins': None, 'call_pre_covered': None,
            'call_p2_attempts':  None, 'call_p2_connects':  None, 'call_p2_talk_mins':  None, 'call_p2_covered':  None,
            'call_p7_attempts':  None, 'call_p7_connects':  None, 'call_p7_talk_mins':  None, 'call_p7_covered':  None,
            'call_p14_attempts': None, 'call_p14_connects': None, 'call_p14_talk_mins': None, 'call_p14_covered': None,
            'call_p14p_attempts':None, 'call_p14p_connects':None, 'call_p14p_talk_mins':None, 'call_p14p_covered':None,
        })
        f_email      = ex.submit(_safe_query, _query_emails_india,     src_client, date_literal, wt_param, label=f'emails/{ev.event_id}',     default={'email_sent': None, 'email_delivered': None, 'email_opened': None, 'email_clicked': None})
        f_sales      = ex.submit(_safe_query, _query_sales_india,      src_client, date_literal, wt_param, label=f'sales/{ev.event_id}',      default={'sales': None, 'revenue': None, 'paid_revenue': None})

        cohort     = f_cohort.result()
        attendance = f_attendance.result()
        call_data  = f_calls.result()
        email      = f_email.result()
        sales_data = f_sales.result()

    snap_row = _cumulative_snapshot_row(ev.event_id, daily, ev.live_at, now, now_iso,
                                        cohort=cohort, attendance=attendance,
                                        call_data=call_data, email=email, sales_data=sales_data)
    return daily_rows, snap_row

def _safe_query(fn, *args, label, default, **kw):
    """Run a query function; on failure, log and return defaults so the cron continues."""
    try:
        return fn(*args, **kw)
    except Exception as e:
        print(f'  [snapshot/bq] {label} failed: {e}')
        return default

def _query_cohort_india(src_client, date_literal, wt_param):
    """Aggregate qualified IQLs by role category + work_ex bucket.
    Maps source 5-category role split into our 4-bucket schema (role_fe is always
    None — source 'Software Engineer' bucket already includes Front-end/Full-stack).
    we_0_2 is always 0 because the source query filters out 0-2 / 3-4 / student leads."""
    from google.cloud import bigquery
    q = f"""
SELECT
  CASE
    WHEN role_domain IN ('Data Science','Data Engineer','Machine Learning / AI','ML / AI','Data Engineer / Data Scientist','Business Intelligence Analyst','Data Analyst / Business Analyst','Data','Machine Learning/DeepLearning','Data Engineering') THEN 'Data'
    WHEN role_domain IS NULL OR role_domain IN ('No / Little coding experience','No Coding Experience','None of the above') THEN 'Null'
    WHEN role_domain IN ('Full Stack','Back-end','Other Software Engineers','iOS Developer','Android Developer','iOS / Android Developer','Front-end','Test Engineer / SDET / QE','QA / Testing','Software Engineer','Front-end / Full stack','Software Engineering','Software Engineering (Frontend, Fullstack, Backend, Test)','Mobile Engineering (iOS/Android)','Core Engineering') THEN 'Software Engineer'
    WHEN role_domain IN ('Engineering Manager - any domain','Product Manager (Tech)','Technical Program Manager','Engineering Manager / Director of Engineering','Project Manager / Product Manager','Engineering Manager','Growth Product Manager','Product Marketing Manager','Tech Product Manager','Management') THEN 'Management'
    WHEN role_domain IN ('SRE / DevOps','Cyber Security','Embedded Software Engineer','Cloud Engineer','Application Packaging Engineer','AWS Cloud Solutions Architect','Cyber Security/Security Engineering','Embedded Systems','DevOps Engineer','Site Reliability Engineer','Site Reliability Engineering') THEN 'Systems'
    ELSE 'Other'
  END AS category,
  CASE
    WHEN work_ex IN ('0-2','0-5','3-4') THEN 'a05'
    WHEN work_ex IN ('5-8','5-10')      THEN 'b510'
    WHEN work_ex IN ('9-15','10-15')    THEN 'c1015'
    WHEN work_ex IN ('16-20','15-20','15+') THEN 'd1520'
    WHEN work_ex = '20+'               THEN 'e20p'
    ELSE 'other'
  END AS we_bucket,
  COUNT(*) AS cnt
FROM (
  SELECT role_domain, work_ex,
    DATE(event_start_date_time, "Asia/Kolkata") AS web_scheduled_date,
    dupe_flag, gql_flag, webinar_type, dupe_logic,
    ROW_NUMBER() OVER (
      PARTITION BY hubspot_ID, DATE(event_start_date_time, "Asia/Kolkata")
      ORDER BY formatted_date ASC
    ) AS rnk
  FROM `ik-marketing-data.India_Leads.US_Domain_combined_view`
  WHERE dupe_logic = 1
)
WHERE rnk = 1
  AND web_scheduled_date IN ({date_literal})
  AND UPPER(webinar_type) = UPPER(@wt)
  AND dupe_flag = 0 AND gql_flag = 0
  AND LOWER(work_ex) NOT LIKE '%student%' AND work_ex NOT IN ('0-2','3-4')
GROUP BY 1, 2"""
    rows = list(src_client.query(q, job_config=bigquery.QueryJobConfig(query_parameters=wt_param)).result())

    out = {'role_sde': 0, 'role_ml': 0, 'role_fe': None,
           'role_management': 0, 'role_systems': 0, 'role_null': 0, 'role_other': 0,
           'we_0_2': 0, 'we_3_5': 0, 'we_6_10': 0,
           'we_10_15': 0, 'we_15_20': 0, 'we_20p': 0, 'we_other': 0, 'we_10p': 0}
    # Map source labels → schema columns. role_ml stores "Data"; we_3_5 stores "0-5".
    # Names kept for back-compat — see CONTEXT.md.
    role_map = {
        'Software Engineer': 'role_sde',
        'Data':              'role_ml',
        'Management':        'role_management',
        'Systems':           'role_systems',
        'Null':              'role_null',
        'Other':             'role_other',
    }
    we_map = {
        'a05':   'we_3_5',
        'b510':  'we_6_10',
        'c1015': 'we_10_15',
        'd1520': 'we_15_20',
        'e20p':  'we_20p',
        'other': 'we_other',
    }
    for r in rows:
        cnt = int(r['cnt'] or 0)
        if (col := role_map.get(r['category'])): out[col] += cnt
        if (col := we_map.get(r['we_bucket'])):  out[col] += cnt
        if r['we_bucket'] in ('c1015', 'd1520', 'e20p'): out['we_10p'] += cnt
    return out

def _query_channel_attendance(event):
    """Per-channel attendance for an aired event, computed at read-time.
    Joins Zoom attendees to lead channel/utm, returns one entry per UI bucket.

    Returns:
      India: {'Meta': {'leads': N, 'attendees': M}, 'CRM': {...}, 'Others': {...}}
      US:    {'YT': {...}, 'Social': {...}, 'L10X Email': {...}, 'L10X Bot': {...},
              'NI Base': {...}, 'Other': {...}}
    Returns None on error / unsupported country."""
    from google.cloud import bigquery
    country = (event.get('country') or '').lower()
    wt      = event.get('webinar_type')
    if not wt:
        return None

    # Build (PT or IST) date list from live_at + day2_live_at
    tz_name = 'America/Los_Angeles' if country in ('us', 'usa') else 'Asia/Kolkata'
    tz_obj  = zoneinfo.ZoneInfo(tz_name)
    dates   = []
    for d in (event.get('live_at'), event.get('day2_live_at')):
        if d is None: continue
        if isinstance(d, str):
            d = datetime.datetime.fromisoformat(d.replace('Z', '+00:00'))
        if isinstance(d, datetime.datetime):
            aware = d if d.tzinfo else d.replace(tzinfo=datetime.timezone.utc)
            dates.append(aware.astimezone(tz_obj).date())
        else:
            dates.append(d)
    if not dates:
        return None
    date_literal = ', '.join(f"DATE '{d.isoformat()}'" for d in dates)
    wt_param     = [bigquery.ScalarQueryParameter('wt', 'STRING', wt)]
    src_client   = bigquery.Client(project='ik-marketing-data')

    if country in ('us', 'usa'):
        q = f"""
WITH attendance AS (
  SELECT CAST(hubspot_id AS INT64) AS hubspot_id
  FROM `ik-marketing-data.Webinar_analytics.webinar_attendee_data_from_zoom`
  WHERE DATE(webinar_start_time, "{tz_name}") IN ({date_literal})
    AND hubspot_id IS NOT NULL
  GROUP BY hubspot_id
),
base AS (
  SELECT leads_hubspot_id, utm_campaign, formatted_date, dupe_flag, gql_flag,
    DATE(event_start_date_time, "{tz_name}") AS web_scheduled_date,
    CASE
      WHEN LOWER(utm_campaign) LIKE "%l10x_social%"               THEN "Social"
      WHEN REGEXP_CONTAINS(LOWER(utm_campaign), r"__l10x__")      THEN "L10X_Email"
      WHEN REGEXP_CONTAINS(LOWER(utm_campaign), r"youtube_l10x")  THEN "YT"
      WHEN LOWER(utm_campaign) LIKE "%nibucket%"                  THEN "NI_Base"
      WHEN LOWER(utm_campaign) LIKE "%l10x-bot%"                  THEN "L10X_Bot"
      ELSE "Other"
    END AS channel_bucket,
    ROW_NUMBER() OVER (
      PARTITION BY leads_hubspot_id, DATE(event_start_date_time, "{tz_name}")
      ORDER BY formatted_date ASC
    ) AS rnk
  FROM `ik-marketing-data.Marketing_data_new_logic.Bq_data_Alumni`
  WHERE UPPER(webinar_type) = UPPER(@wt) AND dupe_logic = 1
)
SELECT channel_bucket, COUNT(*) AS leads,
       COUNTIF(a.hubspot_id IS NOT NULL) AS attendees
FROM (
  SELECT * FROM base
  WHERE rnk = 1
    AND web_scheduled_date IN ({date_literal})
    AND dupe_flag = 0 AND gql_flag = 0
) l
LEFT JOIN attendance a
  ON CAST(l.leads_hubspot_id AS INT64) = a.hubspot_id
GROUP BY channel_bucket"""
        rows = list(src_client.query(q, job_config=bigquery.QueryJobConfig(query_parameters=wt_param)).result())
        # Map source channel_bucket → UI bucket label
        label_map = {
            'YT':         'YT',
            'Social':     'Social',
            'L10X_Email': 'L10X Email',
            'L10X_Bot':   'L10X Bot',
            'NI_Base':    'NI Base',
            'Other':      'Other',
        }
        out = {v: {'leads': 0, 'attendees': 0} for v in label_map.values()}
        for r in rows:
            label = label_map.get(r['channel_bucket'], 'Other')
            out[label]['leads']     += int(r['leads'] or 0)
            out[label]['attendees'] += int(r['attendees'] or 0)
        return out

    if country in ('india', 'in'):
        q = f"""
WITH attendance AS (
  SELECT CAST(hubspot_id AS INT64) AS hubspot_id
  FROM `ik-marketing-data.Webinar_analytics.webinar_attendee_data_from_zoom`
  WHERE DATE(webinar_start_time, "{tz_name}") IN ({date_literal})
    AND hubspot_id IS NOT NULL
  GROUP BY hubspot_id
),
leads AS (
  SELECT CAST(hubspot_ID AS INT64) AS hubspot_id, Channel
  FROM (
    SELECT hubspot_ID, Channel, formatted_date,
      DATE(event_start_date_time, "{tz_name}") AS web_scheduled_date,
      dupe_flag, gql_flag, work_ex, webinar_type, dupe_logic,
      ROW_NUMBER() OVER (
        PARTITION BY hubspot_ID, DATE(event_start_date_time, "{tz_name}")
        ORDER BY formatted_date ASC
      ) AS rnk
    FROM `ik-marketing-data.India_Leads.US_Domain_combined_view`
    WHERE dupe_logic = 1
  )
  WHERE rnk = 1
    AND web_scheduled_date IN ({date_literal})
    AND UPPER(webinar_type) = UPPER(@wt)
    AND dupe_flag = 0 AND gql_flag = 0
    AND LOWER(work_ex) NOT LIKE '%student%' AND work_ex NOT IN ('0-2','3-4')
  QUALIFY ROW_NUMBER() OVER (PARTITION BY hubspot_ID ORDER BY formatted_date ASC) = 1
)
SELECT Channel, COUNT(*) AS leads,
       COUNTIF(a.hubspot_id IS NOT NULL) AS attendees
FROM leads l LEFT JOIN attendance a ON l.hubspot_id = a.hubspot_id
GROUP BY Channel"""
        rows = list(src_client.query(q, job_config=bigquery.QueryJobConfig(query_parameters=wt_param)).result())
        # India: bucket by regex into Meta / CRM / Others (matches snapshot logic)
        out = {'Meta': {'leads': 0, 'attendees': 0},
               'CRM':  {'leads': 0, 'attendees': 0},
               'Others': {'leads': 0, 'attendees': 0}}
        for r in rows:
            ch = (r['Channel'] or '')
            if   _RE_META.search(ch): bucket = 'Meta'
            elif _RE_CRM.search(ch):  bucket = 'CRM'
            else:                     bucket = 'Others'
            out[bucket]['leads']     += int(r['leads'] or 0)
            out[bucket]['attendees'] += int(r['attendees'] or 0)
        return out

    return None

def _query_attendance_india(src_client, date_literal, wt_param):
    """Returns {attendees, attendance_pct}. attendees = unique IQLs in Zoom roster."""
    from google.cloud import bigquery
    q = f"""
WITH attendance AS (
  SELECT CAST(hubspot_id AS INT64) AS hubspot_id
  FROM `ik-marketing-data.Webinar_analytics.webinar_attendee_data_from_zoom`
  WHERE DATE(webinar_start_time, "Asia/Kolkata") IN ({date_literal})
    AND hubspot_id IS NOT NULL
  GROUP BY hubspot_id
),
leads AS (
  SELECT CAST(hubspot_ID AS INT64) AS hubspot_id
  FROM (
    SELECT hubspot_ID, formatted_date,
      DATE(event_start_date_time, "Asia/Kolkata") AS web_scheduled_date,
      dupe_flag, gql_flag, work_ex, webinar_type, dupe_logic,
      ROW_NUMBER() OVER (
        PARTITION BY hubspot_ID, DATE(event_start_date_time, "Asia/Kolkata")
        ORDER BY formatted_date ASC
      ) AS rnk
    FROM `ik-marketing-data.India_Leads.US_Domain_combined_view`
    WHERE dupe_logic = 1
  )
  WHERE rnk = 1
    AND web_scheduled_date IN ({date_literal})
    AND UPPER(webinar_type) = UPPER(@wt)
    AND dupe_flag = 0 AND gql_flag = 0
    AND LOWER(work_ex) NOT LIKE '%student%' AND work_ex NOT IN ('0-2','3-4')
  QUALIFY ROW_NUMBER() OVER (PARTITION BY hubspot_ID ORDER BY formatted_date ASC) = 1
)
SELECT
  COUNT(*) AS total_leads,
  COUNTIF(a.hubspot_id IS NOT NULL) AS attendees
FROM leads l
LEFT JOIN attendance a ON l.hubspot_id = a.hubspot_id"""
    r = list(src_client.query(q, job_config=bigquery.QueryJobConfig(query_parameters=wt_param)).result())
    if not r:
        return {'attendees': None, 'attendance_pct': None}
    total    = int(r[0]['total_leads'] or 0)
    attended = int(r[0]['attendees'] or 0)
    pct      = (100.0 * attended / total) if total > 0 else None
    return {'attendees': attended, 'attendance_pct': pct}

def _query_calls_india(src_client, date_literal, wt_param):
    """Returns pre-webinar top-level metrics for back-compat + extras.call_buckets
    with the 5 lifecycle buckets (pre, 0-2D, 0-7D, 0-14D, 14D+).

    Each bucket carries: attempts, connects, talk_mins, covered_leads (# of distinct
    leads who got ≥1 call in that bucket). ALL post-webinar buckets are CUMULATIVE
    from the webinar date: 0-2D ⊂ 0-7D ⊂ 0-14D ⊂ 14D+ (14D+ = ALL post-webinar calls).
    """
    from google.cloud import bigquery
    q = f"""
WITH base AS (
  SELECT hubspot_ID AS leads_hubspot_id,
    DATETIME(event_start_date_time, 'Asia/Kolkata') AS webinar_start_datetime_ch
  FROM (
    SELECT hubspot_ID, event_start_date_time, formatted_date,
      DATE(event_start_date_time, "Asia/Kolkata") AS web_scheduled_date,
      dupe_flag, gql_flag, webinar_type, dupe_logic, work_ex,
      ROW_NUMBER() OVER (
        PARTITION BY hubspot_ID, DATE(event_start_date_time, "Asia/Kolkata")
        ORDER BY formatted_date ASC
      ) AS rnk
    FROM `ik-marketing-data.India_Leads.US_Domain_combined_view`
    WHERE dupe_logic = 1
  )
  WHERE rnk = 1
    AND web_scheduled_date IN ({date_literal})
    AND UPPER(webinar_type) = UPPER(@wt)
    AND dupe_flag = 0 AND gql_flag = 0
    AND LOWER(work_ex) NOT LIKE '%student%' AND work_ex NOT IN ('0-2','3-4')
),
base_call AS (
  SELECT activity_datetime, DATE(activity_datetime) AS activity_date, hubspot_id, call_duration FROM (
    SELECT
      CASE WHEN EXTRACT(MONTH FROM DATE(DATETIME(timestamp, "Asia/Kolkata"))) IN (11,12,1,2,3)
           THEN DATETIME(timestamp, "Asia/Kolkata") + INTERVAL 13 HOUR + INTERVAL 30 MINUTE
           ELSE DATETIME(timestamp, "Asia/Kolkata") + INTERVAL 12 HOUR + INTERVAL 30 MINUTE
      END AS activity_datetime,
      hubspot_id,
      duration AS call_duration,
      ROW_NUMBER() OVER (PARTITION BY call_id ORDER BY
        CASE WHEN EXTRACT(MONTH FROM DATE(DATETIME(timestamp, "Asia/Kolkata"))) IN (11,12,1,2,3)
             THEN DATETIME(timestamp, "Asia/Kolkata") + INTERVAL 13 HOUR + INTERVAL 30 MINUTE
             ELSE DATETIME(timestamp, "Asia/Kolkata") + INTERVAL 12 HOUR + INTERVAL 30 MINUTE
        END DESC) AS rn
    FROM `ik-marketing-data.Marketing_data_new_logic.call_metadata`
    WHERE DATETIME(timestamp, "Asia/Kolkata") >= DATETIME '2024-07-01'
      AND hubspot_id IN (SELECT leads_hubspot_id FROM base)
  ) WHERE rn = 1
),
final_merge AS (
  SELECT b.leads_hubspot_id, b.webinar_start_datetime_ch,
         bc.activity_datetime, bc.activity_date, bc.call_duration
  FROM base b
  LEFT JOIN base_call bc ON b.leads_hubspot_id = bc.hubspot_id
),
per_lead AS (
  SELECT
    leads_hubspot_id,
    -- Pre-webinar
    COUNT(CASE WHEN activity_datetime < webinar_start_datetime_ch THEN 1 END) AS pre_calls,
    COUNT(CASE WHEN activity_datetime < webinar_start_datetime_ch AND call_duration > 120 THEN 1 END) AS pre_conn,
    SUM(CASE WHEN activity_datetime < webinar_start_datetime_ch THEN call_duration / 60.0 END) AS pre_talk_mins,
    -- Post 0-2D (cumulative window from event date)
    COUNT(DISTINCT CASE WHEN activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 2 DAY) THEN activity_datetime END) AS p2_calls,
    COUNT(DISTINCT CASE WHEN call_duration > 120 AND activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 2 DAY) THEN activity_datetime END) AS p2_conn,
    SUM(CASE WHEN call_duration > 120 AND activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 2 DAY) THEN call_duration / 60.0 ELSE 0 END) AS p2_talk_mins,
    -- Post 0-7D
    COUNT(DISTINCT CASE WHEN activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 7 DAY) THEN activity_datetime END) AS p7_calls,
    COUNT(DISTINCT CASE WHEN call_duration > 120 AND activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 7 DAY) THEN activity_datetime END) AS p7_conn,
    SUM(CASE WHEN call_duration > 120 AND activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 7 DAY) THEN call_duration / 60.0 ELSE 0 END) AS p7_talk_mins,
    -- Post 0-14D
    COUNT(DISTINCT CASE WHEN activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 14 DAY) THEN activity_datetime END) AS p14_calls,
    COUNT(DISTINCT CASE WHEN call_duration > 120 AND activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 14 DAY) THEN activity_datetime END) AS p14_conn,
    SUM(CASE WHEN call_duration > 120 AND activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 14 DAY) THEN call_duration / 60.0 ELSE 0 END) AS p14_talk_mins,
    -- 14D+ (CUMULATIVE: all post-webinar calls, including 0-2D, 0-7D, 0-14D)
    COUNT(DISTINCT CASE WHEN activity_date >= DATE(webinar_start_datetime_ch) THEN activity_datetime END) AS p14p_calls,
    COUNT(DISTINCT CASE WHEN call_duration > 120 AND activity_date >= DATE(webinar_start_datetime_ch) THEN activity_datetime END) AS p14p_conn,
    SUM(CASE WHEN call_duration > 120 AND activity_date >= DATE(webinar_start_datetime_ch) THEN call_duration / 60.0 ELSE 0 END) AS p14p_talk_mins
  FROM final_merge
  GROUP BY leads_hubspot_id
)
SELECT
  COUNT(*) AS total_leads,
  SUM(pre_calls)  AS pre_total_calls,  SUM(pre_conn)  AS pre_total_conn,  SUM(pre_talk_mins)  AS pre_total_talk_mins,  COUNTIF(pre_calls > 0)  AS pre_covered_leads,
  SUM(p2_calls)   AS p2_total_calls,   SUM(p2_conn)   AS p2_total_conn,   SUM(p2_talk_mins)   AS p2_total_talk_mins,   COUNTIF(p2_calls > 0)   AS p2_covered_leads,
  SUM(p7_calls)   AS p7_total_calls,   SUM(p7_conn)   AS p7_total_conn,   SUM(p7_talk_mins)   AS p7_total_talk_mins,   COUNTIF(p7_calls > 0)   AS p7_covered_leads,
  SUM(p14_calls)  AS p14_total_calls,  SUM(p14_conn)  AS p14_total_conn,  SUM(p14_talk_mins)  AS p14_total_talk_mins,  COUNTIF(p14_calls > 0)  AS p14_covered_leads,
  SUM(p14p_calls) AS p14p_total_calls, SUM(p14p_conn) AS p14p_total_conn, SUM(p14p_talk_mins) AS p14p_total_talk_mins, COUNTIF(p14p_calls > 0) AS p14p_covered_leads
FROM per_lead"""
    r = list(src_client.query(q, job_config=bigquery.QueryJobConfig(query_parameters=wt_param)).result())
    empty = {
        'calls_attempted': None, 'calls_connected': None, 'avg_talk_seconds': None,
        'call_total_leads': None,
        'call_pre_attempts': None, 'call_pre_connects': None, 'call_pre_talk_mins': None, 'call_pre_covered': None,
        'call_p2_attempts':  None, 'call_p2_connects':  None, 'call_p2_talk_mins':  None, 'call_p2_covered':  None,
        'call_p7_attempts':  None, 'call_p7_connects':  None, 'call_p7_talk_mins':  None, 'call_p7_covered':  None,
        'call_p14_attempts': None, 'call_p14_connects': None, 'call_p14_talk_mins': None, 'call_p14_covered': None,
        'call_p14p_attempts':None, 'call_p14p_connects':None, 'call_p14p_talk_mins':None, 'call_p14p_covered':None,
    }
    if not r:
        return empty
    row = r[0]
    n = lambda k: int(row[k] or 0)
    g = lambda k: float(row[k] or 0)
    pre_conn = n('pre_total_conn')
    pre_talk_mins = g('pre_total_talk_mins')
    avg_talk_sec = (pre_talk_mins * 60.0 / pre_conn) if pre_conn > 0 else None
    # Write lifecycle stages to TOP-LEVEL columns the event page actually reads.
    # The earlier `extras` JSON dict was rejected by BQ streaming insert because
    # the table's `extras` column is typed STRUCT, not JSON — so every event's
    # snapshot row was silently failing to insert. Top-level columns sidestep that.
    return {
        # Back-compat top-level (pre-webinar values; existing semantics)
        'calls_attempted':  n('pre_total_calls'),
        'calls_connected':  pre_conn,
        'avg_talk_seconds': avg_talk_sec,
        # Lifecycle stages (consumed by event.html Call Efforts dashboard)
        'call_total_leads':    n('total_leads'),
        'call_pre_attempts':   n('pre_total_calls'),
        'call_pre_connects':   n('pre_total_conn'),
        'call_pre_talk_mins':  g('pre_total_talk_mins'),
        'call_pre_covered':    n('pre_covered_leads'),
        'call_p2_attempts':    n('p2_total_calls'),
        'call_p2_connects':    n('p2_total_conn'),
        'call_p2_talk_mins':   g('p2_total_talk_mins'),
        'call_p2_covered':     n('p2_covered_leads'),
        'call_p7_attempts':    n('p7_total_calls'),
        'call_p7_connects':    n('p7_total_conn'),
        'call_p7_talk_mins':   g('p7_total_talk_mins'),
        'call_p7_covered':     n('p7_covered_leads'),
        'call_p14_attempts':   n('p14_total_calls'),
        'call_p14_connects':   n('p14_total_conn'),
        'call_p14_talk_mins':  g('p14_total_talk_mins'),
        'call_p14_covered':    n('p14_covered_leads'),
        'call_p14p_attempts':  n('p14p_total_calls'),
        'call_p14p_connects':  n('p14p_total_conn'),
        'call_p14p_talk_mins': g('p14p_total_talk_mins'),
        'call_p14p_covered':   n('p14p_covered_leads'),
    }

def _query_sales_india(src_client, date_literal, wt_param):
    """Returns sales count, total INR revenue, paid (Meta/Facebook) INR revenue.
    Converts net_revenue from USD to INR per-sale using each sale's month avg
    USD/INR rate from events.fx_rates_monthly. Falls back to 84 if rate missing."""
    from google.cloud import bigquery
    q = f"""
WITH sales AS (
  SELECT Sale_date, net_revenue, Channel, dupe_flag,
    FORMAT_DATE('%Y-%m', Sale_date) AS sale_month
  FROM (
    SELECT Sale_date, net_revenue, Channel, dupe_flag, webinar_type, work_ex,
      DATE(event_start_date_time, "Asia/Kolkata") AS web_scheduled_date,
      ROW_NUMBER() OVER (
        PARTITION BY hubspot_ID, DATE(event_start_date_time, "Asia/Kolkata")
        ORDER BY formatted_date ASC
      ) AS rnk
    FROM `ik-marketing-data.India_Leads.US_Domain_combined_view`
    WHERE dupe_logic = 1
      AND COALESCE(Alumni_Stats, 'New_Lead') = 'New_Lead'   -- exclude alumni-repeat sales
  )
  WHERE rnk = 1
    AND web_scheduled_date IN ({date_literal})
    AND UPPER(webinar_type) = UPPER(@wt)
    AND dupe_flag = 0
    -- No work_ex exclusion for sales — count paying customers regardless of experience level.
),
sales_inr AS (
  SELECT s.Sale_date, s.Channel,
    s.net_revenue * COALESCE(fx.rate, 84.0) AS rev_inr
  FROM sales s
  LEFT JOIN `masterclass-automation-ik.events.fx_rates_monthly` fx
    ON fx.month = s.sale_month AND fx.base = 'USD' AND fx.quote = 'INR'
)
SELECT
  COUNTIF(Sale_date IS NOT NULL)                                  AS sales,
  SUM(CASE WHEN Sale_date IS NOT NULL THEN rev_inr END)            AS revenue,
  SUM(CASE WHEN Sale_date IS NOT NULL AND Channel = 'Facebook'
           THEN rev_inr END)                                       AS paid_revenue
FROM sales_inr"""
    r = list(src_client.query(q, job_config=bigquery.QueryJobConfig(query_parameters=wt_param)).result())
    if not r:
        return {'sales': 0, 'revenue': 0.0, 'paid_revenue': 0.0}
    row = r[0]
    return {
        'sales':        int(row['sales'] or 0),
        'revenue':      float(row['revenue'] or 0.0),
        'paid_revenue': float(row['paid_revenue'] or 0.0),
    }

def _query_emails_india(src_client, date_literal, wt_param):
    """Returns {email_sent, email_delivered, email_opened, email_clicked} for reminder campaigns."""
    from google.cloud import bigquery
    q = f"""
WITH registered_leads AS (
  SELECT CAST(hubspot_ID AS INT64) AS hubspot_id
  FROM (
    SELECT hubspot_ID, formatted_date,
      DATE(event_start_date_time, "Asia/Kolkata") AS web_scheduled_date,
      dupe_flag, gql_flag, work_ex, webinar_type, dupe_logic,
      ROW_NUMBER() OVER (
        PARTITION BY hubspot_ID, DATE(event_start_date_time, "Asia/Kolkata")
        ORDER BY formatted_date ASC
      ) AS rnk
    FROM `ik-marketing-data.India_Leads.US_Domain_combined_view`
    WHERE dupe_logic = 1
  )
  WHERE rnk = 1
    AND web_scheduled_date IN ({date_literal})
    AND UPPER(webinar_type) = UPPER(@wt)
    AND dupe_flag = 0 AND gql_flag = 0
    AND LOWER(work_ex) NOT LIKE '%student%' AND work_ex NOT IN ('0-2','3-4')
  QUALIFY ROW_NUMBER() OVER (PARTITION BY hubspot_ID ORDER BY formatted_date ASC) = 1
),
email_sent AS (
  SELECT CAST(hubspot_id AS INT64) AS hubspot_id, email_campaign_id
  FROM `ik-marketing-data.Email.Marketing_Email_Data`
  WHERE CAST(hubspot_id AS INT64) IN (SELECT hubspot_id FROM registered_leads)
    AND DATE(event_timestamp, "Asia/Kolkata") IN ({date_literal})
    AND event_name = 'SENT'
    AND LOWER(email_name) LIKE '%reminder%'
    AND NOT REGEXP_CONTAINS(email_name, r'(?i)_eu_')
  GROUP BY hubspot_id, email_campaign_id
),
delivered_events AS (
  SELECT DISTINCT CAST(hubspot_id AS INT64) AS hubspot_id
  FROM `ik-marketing-data.Email.Marketing_Email_Data`
  WHERE CAST(hubspot_id AS INT64) IN (SELECT hubspot_id FROM email_sent)
    AND event_name = 'DELIVERED'
),
engagement_events AS (
  SELECT CAST(e.hubspot_id AS INT64) AS hubspot_id, s.email_campaign_id, e.event_name
  FROM `ik-marketing-data.Email.Marketing_Email_Data` e
  INNER JOIN email_sent s
    ON CAST(e.hubspot_id AS INT64) = s.hubspot_id AND e.email_campaign_id = s.email_campaign_id
  WHERE e.event_name IN ('OPEN','CLICK')
),
funnel AS (
  SELECT
    s.hubspot_id, s.email_campaign_id,
    1 AS is_sent,
    MAX(CASE
      WHEN d.hubspot_id IS NOT NULL THEN 1
      WHEN eng.event_name IN ('OPEN','CLICK') THEN 1
      ELSE 0
    END) AS is_delivered,
    MAX(CASE WHEN eng.event_name IN ('OPEN','CLICK') THEN 1 ELSE 0 END) AS is_opened,
    MAX(CASE WHEN eng.event_name = 'CLICK' THEN 1 ELSE 0 END) AS is_clicked
  FROM email_sent s
  LEFT JOIN delivered_events d ON s.hubspot_id = d.hubspot_id
  LEFT JOIN engagement_events eng ON s.hubspot_id = eng.hubspot_id AND s.email_campaign_id = eng.email_campaign_id
  GROUP BY s.hubspot_id, s.email_campaign_id
)
SELECT
  COUNT(*)                                 AS sent,
  SUM(is_delivered)                        AS delivered,
  SUM(is_opened)                           AS opened,
  SUM(is_clicked)                          AS clicked
FROM funnel"""
    r = list(src_client.query(q, job_config=bigquery.QueryJobConfig(query_parameters=wt_param)).result())
    if not r:
        return {'email_sent': None, 'email_delivered': None, 'email_opened': None, 'email_clicked': None}
    return {
        'email_sent':      int(r[0]['sent']      or 0),
        'email_delivered': int(r[0]['delivered'] or 0),
        'email_opened':    int(r[0]['opened']    or 0),
        'email_clicked':   int(r[0]['clicked']   or 0),
    }

# ─── USA snapshot path ───────────────────────────────────────────────────────
# Uses Marketing_data_new_logic.Bq_data_Alumni (leads) + Google_Sheets.Combined_Spend_data
# (Meta spend, USD). PT timezone. webinar_type is always MASTERCLASS_EVENT_AI.
# Calls + email funnel are not wired for US — legacy tool didn't have them either.

_PT_TZ = zoneinfo.ZoneInfo('America/Los_Angeles')

def _build_event_snapshot_us(client, ev, now, now_iso):
    from google.cloud import bigquery
    live_dates = []
    for d in [ev.live_at, ev.day2_live_at]:
        if d is None: continue
        if isinstance(d, datetime.datetime):
            aware = d if d.tzinfo else d.replace(tzinfo=datetime.timezone.utc)
            live_dates.append(aware.astimezone(_PT_TZ).date())
        else:
            live_dates.append(d)
    if not live_dates:
        return [], _cumulative_snapshot_row(ev.event_id, {}, ev.live_at, now, now_iso)

    date_literal = ', '.join(f"DATE '{d.isoformat()}'" for d in live_dates)
    src_client   = bigquery.Client(project='ik-marketing-data')
    wt_param     = [bigquery.ScalarQueryParameter('wt', 'STRING', ev.webinar_type)]

    # ─── 1. Leads + spends — qualified (dupe_flag=0, gql_flag=0) ──────────────
    query = f"""
WITH base AS (
  SELECT
    formatted_date, utm_campaign,
    DATE(event_start_date_time, "America/Los_Angeles") AS web_scheduled_date,
    CASE
      WHEN LOWER(utm_campaign) LIKE "%l10x_social%"               THEN "Social"
      WHEN REGEXP_CONTAINS(LOWER(utm_campaign), r"__l10x__")      THEN "L10X_Email"
      WHEN REGEXP_CONTAINS(LOWER(utm_campaign), r"youtube_l10x")  THEN "YT"
      WHEN LOWER(utm_campaign) LIKE "%nibucket%"                  THEN "NI_Base"
      WHEN LOWER(utm_campaign) LIKE "%l10x-bot%"                  THEN "L10X_Bot"
      ELSE "Other"
    END AS channel_bucket,
    leads_hubspot_id, dupe_flag, gql_flag, dupe_logic,
    ROW_NUMBER() OVER (
      PARTITION BY leads_hubspot_id, DATE(event_start_date_time, "America/Los_Angeles")
      ORDER BY formatted_date ASC
    ) AS rnk
  FROM `ik-marketing-data.Marketing_data_new_logic.Bq_data_Alumni`
  WHERE UPPER(webinar_type) = UPPER(@wt) AND dupe_logic = 1
),
leads AS (
  SELECT DATE(formatted_date) AS registration_date, channel_bucket, COUNT(*) AS total_leads
  FROM base
  WHERE rnk = 1
    AND web_scheduled_date IN ({date_literal})
    AND dupe_flag = 0 AND gql_flag = 0
  GROUP BY registration_date, channel_bucket
),
spends AS (
  -- Dedup exact-duplicate rows before summing (same fix as India — the source
  -- ETL can double-insert; row unique-key is (date, campaign, ad_group, device)).
  SELECT spend_date, SUM(cost_usd_) AS total_spend
  FROM (
    SELECT DATE(campaign_date, "America/Los_Angeles") AS spend_date,
           campaign_name, ad_group_id, device,
           ANY_VALUE(cost_usd_) AS cost_usd_
    FROM `ik-marketing-data.Google_Sheets.Combined_Spend_data`
    -- Widened 2026-06-22: keep event-specific paid campaigns whose names contain
    -- 'masterclass' but not l10x/meta/facebook (e.g., Pilot_Taboola_US_46_Masterclass_AI_Reinvent).
    -- Then EXCLUDE non-event-specific evergreen/promo campaigns:
    --   'recorded_masterclass' (Performance_Max_*, Quora_*_Recorded_Masterclass_*) — evergreen lead-gen
    --   'holiday_offer'        (L10X_Meta_US_Holiday_Offer, _New, _Canada_*)        — promo, not tied to an event
    WHERE (LOWER(campaign_name) LIKE "%l10x%" OR LOWER(campaign_name) LIKE "%masterclass%")
      AND LOWER(campaign_name) NOT LIKE "%recorded_masterclass%"
      AND LOWER(campaign_name) NOT LIKE "%holiday_offer%"
    GROUP BY spend_date, campaign_name, ad_group_id, device
  )
  GROUP BY spend_date
)
SELECT l.registration_date, l.channel_bucket, l.total_leads, s.total_spend
FROM leads l LEFT JOIN spends s ON l.registration_date = s.spend_date
ORDER BY l.registration_date, l.channel_bucket"""

    from google.cloud import bigquery as _bq
    rows = list(src_client.query(query, job_config=_bq.QueryJobConfig(query_parameters=wt_param)).result())

    # US channel → schema bucket mapping:
    #   meta_regs   ← Social + YT          (paid acquisition)
    #   crm_regs    ← L10X_Email + L10X_Bot
    #   other_regs  ← NI_Base + Other
    # meta_spend is total US spend per date (the source query already filters to
    # meta-flavored spend rows). Spend is in USD, not INR — the UI shows the ₹ symbol
    # via fmtMoney; a label tweak for US is a follow-up.
    daily = {}
    spend_seen = set()
    us_totals = {
        'us_yt_regs': 0, 'us_social_regs': 0,
        'us_l10x_email_regs': 0, 'us_l10x_bot_regs': 0,
        'us_ni_base_regs': 0, 'us_other_regs': 0,
    }
    for r in rows:
        d = r['registration_date'].isoformat() if r['registration_date'] else None
        if not d: continue
        ch    = r['channel_bucket'] or ''
        count = float(r['total_leads'] or 0)
        spend = float(r['total_spend'] or 0)

        b = daily.setdefault(d, {
            'meta_regs': 0.0, 'meta_spend': 0.0,
            'crm_regs': 0.0,
            'other_regs': 0.0, 'other_spend': 0.0,
        })
        if spend and d not in spend_seen:
            spend_seen.add(d)
            b['meta_spend'] += spend
        # Existing 3-bucket mapping into India-shaped schema (kept for back-compat
        # with the daily curve renderer + listing query).
        if   ch in ('Social', 'YT'):           b['meta_regs']  += count
        elif ch in ('L10X_Email', 'L10X_Bot'): b['crm_regs']   += count
        else:                                  b['other_regs'] += count
        # US-specific 6-bucket totals (used by the event-page funnel for US events).
        col = {'YT': 'us_yt_regs', 'Social': 'us_social_regs',
               'L10X_Email': 'us_l10x_email_regs', 'L10X_Bot': 'us_l10x_bot_regs',
               'NI_Base': 'us_ni_base_regs'}.get(ch, 'us_other_regs')
        us_totals[col] += int(count)

    daily_rows = []
    for d, b in sorted(daily.items()):
        total_regs  = b['meta_regs'] + b['crm_regs'] + b['other_regs']
        total_spend = b['meta_spend'] + b['other_spend']
        cpiql       = (b['meta_spend'] / b['meta_regs']) if b['meta_regs'] > 0 else None
        daily_rows.append({
            'event_id':          ev.event_id,
            'registration_date': d,
            'snapshot_at':       now_iso,
            'meta_regs':         int(b['meta_regs']),
            'meta_spend':        b['meta_spend'],
            'crm_regs':          int(b['crm_regs']),
            'other_regs':        int(b['other_regs']),
            'other_spend':       b['other_spend'],
            'total_regs':        int(total_regs),
            'total_spend':       total_spend,
            'cpiql':             cpiql,
            'extras':            None,
        })

    # ─── 2-5. Cohort + Attendance + Calls + Sales in parallel ───────────────────
    cohort_default = {'role_sde': None, 'role_ml': None, 'role_fe': None,
                      'role_management': None, 'role_systems': None, 'role_null': None, 'role_other': None,
                      'we_0_2': None, 'we_3_5': None, 'we_6_10': None,
                      'we_10_15': None, 'we_15_20': None, 'we_20p': None,
                      'we_other': None, 'we_10p': None}
    calls_default = {
        'calls_attempted': None, 'calls_connected': None, 'avg_talk_seconds': None,
        'call_total_leads': None,
        'call_pre_attempts': None, 'call_pre_connects': None, 'call_pre_talk_mins': None, 'call_pre_covered': None,
        'call_p2_attempts':  None, 'call_p2_connects':  None, 'call_p2_talk_mins':  None, 'call_p2_covered':  None,
        'call_p7_attempts':  None, 'call_p7_connects':  None, 'call_p7_talk_mins':  None, 'call_p7_covered':  None,
        'call_p14_attempts': None, 'call_p14_connects': None, 'call_p14_talk_mins': None, 'call_p14_covered': None,
        'call_p14p_attempts':None, 'call_p14p_connects':None, 'call_p14p_talk_mins':None, 'call_p14p_covered':None,
    }
    with ThreadPoolExecutor(max_workers=4) as ex:
        f_cohort     = ex.submit(_safe_query, _query_cohort_us,     src_client, date_literal, wt_param, label=f'cohort/{ev.event_id}',     default=cohort_default)
        f_attendance = ex.submit(_safe_query, _query_attendance_us, src_client, date_literal, wt_param, label=f'attendance/{ev.event_id}', default={'attendees': None, 'attendance_pct': None})
        f_calls      = ex.submit(_safe_query, _query_calls_us,      src_client, date_literal, wt_param, label=f'calls/{ev.event_id}',      default=calls_default)
        f_sales      = ex.submit(_safe_query, _query_sales_us,      src_client, date_literal, wt_param, label=f'sales/{ev.event_id}',      default={'sales': None, 'revenue': None, 'paid_revenue': None})
        cohort     = f_cohort.result()
        attendance = f_attendance.result()
        call_data  = f_calls.result()
        sales_data = f_sales.result()

    snap_row = _cumulative_snapshot_row(ev.event_id, daily, ev.live_at, now, now_iso,
                                        cohort=cohort, attendance=attendance,
                                        call_data=call_data, sales_data=sales_data, extra=us_totals)
    return daily_rows, snap_row

def _query_cohort_us(src_client, date_literal, wt_param):
    """US cohort. Same role/work_ex CASE buckets as India, but on Bq_data_Alumni,
    PT timezone, no work_ex student/0-2/3-4 filter (US doesn't filter cohort)."""
    from google.cloud import bigquery
    q = f"""
SELECT
  CASE
    WHEN role_domain IN ('Data Science','Data Engineer','Machine Learning / AI','ML / AI','Data Engineer / Data Scientist','Business Intelligence Analyst','Data Analyst / Business Analyst','Data','Machine Learning/DeepLearning','Data Engineering') THEN 'Data'
    WHEN role_domain IS NULL OR role_domain IN ('No / Little coding experience','No Coding Experience','None of the above') THEN 'Null'
    WHEN role_domain IN ('Full Stack','Back-end','Other Software Engineers','iOS Developer','Android Developer','iOS / Android Developer','Front-end','Test Engineer / SDET / QE','QA / Testing','Software Engineer','Front-end / Full stack','Software Engineering','Software Engineering (Frontend, Fullstack, Backend, Test)','Mobile Engineering (iOS/Android)','Core Engineering') THEN 'Software Engineer'
    WHEN role_domain IN ('Engineering Manager - any domain','Product Manager (Tech)','Technical Program Manager','Engineering Manager / Director of Engineering','Project Manager / Product Manager','Engineering Manager','Growth Product Manager','Product Marketing Manager','Tech Product Manager','Management') THEN 'Management'
    WHEN role_domain IN ('SRE / DevOps','Cyber Security','Embedded Software Engineer','Cloud Engineer','Application Packaging Engineer','AWS Cloud Solutions Architect','Cyber Security/Security Engineering','Embedded Systems','DevOps Engineer','Site Reliability Engineer','Site Reliability Engineering') THEN 'Systems'
    ELSE 'Other'
  END AS category,
  CASE
    WHEN work_ex IN ('0-2','0-5','3-4') THEN 'a05'
    WHEN work_ex IN ('5-8','5-10')      THEN 'b510'
    WHEN work_ex IN ('9-15','10-15')    THEN 'c1015'
    WHEN work_ex IN ('16-20','15-20','15+') THEN 'd1520'
    WHEN work_ex = '20+'               THEN 'e20p'
    ELSE 'other'
  END AS we_bucket,
  COUNT(*) AS cnt
FROM (
  SELECT role_domain, work_ex,
    DATE(event_start_date_time, "America/Los_Angeles") AS web_scheduled_date,
    dupe_flag, gql_flag, dupe_logic,
    ROW_NUMBER() OVER (
      PARTITION BY leads_hubspot_id, DATE(event_start_date_time, "America/Los_Angeles")
      ORDER BY formatted_date ASC
    ) AS rnk
  FROM `ik-marketing-data.Marketing_data_new_logic.Bq_data_Alumni`
  WHERE dupe_logic = 1 AND UPPER(webinar_type) = UPPER(@wt)
)
WHERE rnk = 1
  AND web_scheduled_date IN ({date_literal})
  AND dupe_flag = 0 AND gql_flag = 0
GROUP BY 1, 2"""
    rows = list(src_client.query(q, job_config=bigquery.QueryJobConfig(query_parameters=wt_param)).result())
    out = {'role_sde': 0, 'role_ml': 0, 'role_fe': None,
           'role_management': 0, 'role_systems': 0, 'role_null': 0, 'role_other': 0,
           'we_0_2': 0, 'we_3_5': 0, 'we_6_10': 0,
           'we_10_15': 0, 'we_15_20': 0, 'we_20p': 0, 'we_other': 0, 'we_10p': 0}
    role_map = {'Software Engineer': 'role_sde', 'Data': 'role_ml',
                'Management': 'role_management', 'Systems': 'role_systems',
                'Null': 'role_null', 'Other': 'role_other'}
    we_map = {'a05': 'we_3_5', 'b510': 'we_6_10', 'c1015': 'we_10_15',
              'd1520': 'we_15_20', 'e20p': 'we_20p', 'other': 'we_other'}
    for r in rows:
        cnt = int(r['cnt'] or 0)
        if (col := role_map.get(r['category'])): out[col] += cnt
        if (col := we_map.get(r['we_bucket'])):  out[col] += cnt
        if r['we_bucket'] in ('c1015', 'd1520', 'e20p'): out['we_10p'] += cnt
    return out

def _query_attendance_us(src_client, date_literal, wt_param):
    """US attendance — qualified IQLs from Bq_data_Alumni joined to Zoom roster
    by hubspot_id, filtered to PT date matching live_at."""
    from google.cloud import bigquery
    q = f"""
WITH attendance AS (
  SELECT CAST(hubspot_id AS INT64) AS hubspot_id
  FROM `ik-marketing-data.Webinar_analytics.webinar_attendee_data_from_zoom`
  WHERE DATE(webinar_start_time, "America/Los_Angeles") IN ({date_literal})
    AND hubspot_id IS NOT NULL
  GROUP BY hubspot_id
),
leads AS (
  SELECT CAST(leads_hubspot_id AS INT64) AS hubspot_id
  FROM (
    SELECT leads_hubspot_id, formatted_date,
      DATE(event_start_date_time, "America/Los_Angeles") AS web_scheduled_date,
      dupe_flag, gql_flag, dupe_logic,
      ROW_NUMBER() OVER (
        PARTITION BY leads_hubspot_id, DATE(event_start_date_time, "America/Los_Angeles")
        ORDER BY formatted_date ASC
      ) AS rnk
    FROM `ik-marketing-data.Marketing_data_new_logic.Bq_data_Alumni`
    WHERE dupe_logic = 1 AND UPPER(webinar_type) = UPPER(@wt)
  )
  WHERE rnk = 1
    AND web_scheduled_date IN ({date_literal})
    AND dupe_flag = 0 AND gql_flag = 0
  QUALIFY ROW_NUMBER() OVER (PARTITION BY leads_hubspot_id ORDER BY formatted_date ASC) = 1
)
SELECT COUNT(*) AS total_leads,
       COUNTIF(a.hubspot_id IS NOT NULL) AS attendees
FROM leads l LEFT JOIN attendance a ON l.hubspot_id = a.hubspot_id"""
    r = list(src_client.query(q, job_config=bigquery.QueryJobConfig(query_parameters=wt_param)).result())
    if not r:
        return {'attendees': None, 'attendance_pct': None}
    total    = int(r[0]['total_leads'] or 0)
    attended = int(r[0]['attendees'] or 0)
    pct      = (100.0 * attended / total) if total > 0 else None
    return {'attendees': attended, 'attendance_pct': pct}

def _query_calls_us(src_client, date_literal, wt_param):
    """US call lifecycle. Same shape as India: 5 windows (pre / 0-2D / 0-7D /
    0-14D / 14D+) with attempts, connects (duration>120s), talk_mins, and
    distinct covered leads per window.

    Differences from India:
      - Lead source = `Bq_data_Alumni` (was `US_Domain_combined_view`)
      - PT timezone for event start (was Asia/Kolkata)
      - leads_hubspot_id column (was hubspot_ID)
      - No work_ex / student / 0-2 / 3-4 filter (US doesn't filter cohort)

    Per user (2026-06-22): "use the same India call logic, swap the lead table".
    The +12.5h / +13.5h timestamp adjust on `call_metadata.timestamp` is kept
    as-is, matching India. If US numbers look off this is the first knob —
    see QA_REPORT.md for context on why the shift is suspect."""
    from google.cloud import bigquery
    q = f"""
WITH base AS (
  SELECT leads_hubspot_id,
    DATETIME(event_start_date_time, 'America/Los_Angeles') AS webinar_start_datetime_ch
  FROM (
    SELECT leads_hubspot_id, event_start_date_time, formatted_date,
      DATE(event_start_date_time, "America/Los_Angeles") AS web_scheduled_date,
      dupe_flag, gql_flag, webinar_type, dupe_logic,
      ROW_NUMBER() OVER (
        PARTITION BY leads_hubspot_id, DATE(event_start_date_time, "America/Los_Angeles")
        ORDER BY formatted_date ASC
      ) AS rnk
    FROM `ik-marketing-data.Marketing_data_new_logic.Bq_data_Alumni`
    WHERE dupe_logic = 1
  )
  WHERE rnk = 1
    AND web_scheduled_date IN ({date_literal})
    AND UPPER(webinar_type) = UPPER(@wt)
    AND dupe_flag = 0 AND gql_flag = 0
),
base_call AS (
  SELECT activity_datetime, DATE(activity_datetime) AS activity_date, hubspot_id, call_duration FROM (
    SELECT
      CASE WHEN EXTRACT(MONTH FROM DATE(DATETIME(timestamp, "Asia/Kolkata"))) IN (11,12,1,2,3)
           THEN DATETIME(timestamp, "Asia/Kolkata") + INTERVAL 13 HOUR + INTERVAL 30 MINUTE
           ELSE DATETIME(timestamp, "Asia/Kolkata") + INTERVAL 12 HOUR + INTERVAL 30 MINUTE
      END AS activity_datetime,
      CAST(hubspot_id AS STRING) AS hubspot_id,
      duration AS call_duration,
      ROW_NUMBER() OVER (PARTITION BY call_id ORDER BY
        CASE WHEN EXTRACT(MONTH FROM DATE(DATETIME(timestamp, "Asia/Kolkata"))) IN (11,12,1,2,3)
             THEN DATETIME(timestamp, "Asia/Kolkata") + INTERVAL 13 HOUR + INTERVAL 30 MINUTE
             ELSE DATETIME(timestamp, "Asia/Kolkata") + INTERVAL 12 HOUR + INTERVAL 30 MINUTE
        END DESC) AS rn
    FROM `ik-marketing-data.Marketing_data_new_logic.call_metadata`
    WHERE DATETIME(timestamp, "Asia/Kolkata") >= DATETIME '2024-07-01'
      AND CAST(hubspot_id AS STRING) IN (SELECT CAST(leads_hubspot_id AS STRING) FROM base)
  ) WHERE rn = 1
),
final_merge AS (
  SELECT b.leads_hubspot_id, b.webinar_start_datetime_ch,
         bc.activity_datetime, bc.activity_date, bc.call_duration
  FROM base b
  LEFT JOIN base_call bc ON CAST(b.leads_hubspot_id AS STRING) = bc.hubspot_id
),
per_lead AS (
  SELECT
    leads_hubspot_id,
    COUNT(CASE WHEN activity_datetime < webinar_start_datetime_ch THEN 1 END) AS pre_calls,
    COUNT(CASE WHEN activity_datetime < webinar_start_datetime_ch AND call_duration > 120 THEN 1 END) AS pre_conn,
    SUM(CASE WHEN activity_datetime < webinar_start_datetime_ch THEN call_duration / 60.0 END) AS pre_talk_mins,
    COUNT(DISTINCT CASE WHEN activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 2 DAY) THEN activity_datetime END) AS p2_calls,
    COUNT(DISTINCT CASE WHEN call_duration > 120 AND activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 2 DAY) THEN activity_datetime END) AS p2_conn,
    SUM(CASE WHEN call_duration > 120 AND activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 2 DAY) THEN call_duration / 60.0 ELSE 0 END) AS p2_talk_mins,
    COUNT(DISTINCT CASE WHEN activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 7 DAY) THEN activity_datetime END) AS p7_calls,
    COUNT(DISTINCT CASE WHEN call_duration > 120 AND activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 7 DAY) THEN activity_datetime END) AS p7_conn,
    SUM(CASE WHEN call_duration > 120 AND activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 7 DAY) THEN call_duration / 60.0 ELSE 0 END) AS p7_talk_mins,
    COUNT(DISTINCT CASE WHEN activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 14 DAY) THEN activity_datetime END) AS p14_calls,
    COUNT(DISTINCT CASE WHEN call_duration > 120 AND activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 14 DAY) THEN activity_datetime END) AS p14_conn,
    SUM(CASE WHEN call_duration > 120 AND activity_date BETWEEN DATE(webinar_start_datetime_ch) AND DATE_ADD(DATE(webinar_start_datetime_ch), INTERVAL 14 DAY) THEN call_duration / 60.0 ELSE 0 END) AS p14_talk_mins,
    -- 14D+ (CUMULATIVE: all post-webinar calls, including 0-2D, 0-7D, 0-14D)
    COUNT(DISTINCT CASE WHEN activity_date >= DATE(webinar_start_datetime_ch) THEN activity_datetime END) AS p14p_calls,
    COUNT(DISTINCT CASE WHEN call_duration > 120 AND activity_date >= DATE(webinar_start_datetime_ch) THEN activity_datetime END) AS p14p_conn,
    SUM(CASE WHEN call_duration > 120 AND activity_date >= DATE(webinar_start_datetime_ch) THEN call_duration / 60.0 ELSE 0 END) AS p14p_talk_mins
  FROM final_merge
  GROUP BY leads_hubspot_id
)
SELECT
  COUNT(*) AS total_leads,
  SUM(pre_calls)  AS pre_total_calls,  SUM(pre_conn)  AS pre_total_conn,  SUM(pre_talk_mins)  AS pre_total_talk_mins,  COUNTIF(pre_calls > 0)  AS pre_covered_leads,
  SUM(p2_calls)   AS p2_total_calls,   SUM(p2_conn)   AS p2_total_conn,   SUM(p2_talk_mins)   AS p2_total_talk_mins,   COUNTIF(p2_calls > 0)   AS p2_covered_leads,
  SUM(p7_calls)   AS p7_total_calls,   SUM(p7_conn)   AS p7_total_conn,   SUM(p7_talk_mins)   AS p7_total_talk_mins,   COUNTIF(p7_calls > 0)   AS p7_covered_leads,
  SUM(p14_calls)  AS p14_total_calls,  SUM(p14_conn)  AS p14_total_conn,  SUM(p14_talk_mins)  AS p14_total_talk_mins,  COUNTIF(p14_calls > 0)  AS p14_covered_leads,
  SUM(p14p_calls) AS p14p_total_calls, SUM(p14p_conn) AS p14p_total_conn, SUM(p14p_talk_mins) AS p14p_total_talk_mins, COUNTIF(p14p_calls > 0) AS p14p_covered_leads
FROM per_lead"""
    r = list(src_client.query(q, job_config=bigquery.QueryJobConfig(query_parameters=wt_param)).result())
    empty = {
        'calls_attempted': None, 'calls_connected': None, 'avg_talk_seconds': None,
        'call_total_leads': None,
        'call_pre_attempts': None, 'call_pre_connects': None, 'call_pre_talk_mins': None, 'call_pre_covered': None,
        'call_p2_attempts':  None, 'call_p2_connects':  None, 'call_p2_talk_mins':  None, 'call_p2_covered':  None,
        'call_p7_attempts':  None, 'call_p7_connects':  None, 'call_p7_talk_mins':  None, 'call_p7_covered':  None,
        'call_p14_attempts': None, 'call_p14_connects': None, 'call_p14_talk_mins': None, 'call_p14_covered': None,
        'call_p14p_attempts':None, 'call_p14p_connects':None, 'call_p14p_talk_mins':None, 'call_p14p_covered':None,
    }
    if not r:
        return empty
    row = r[0]
    n = lambda k: int(row[k] or 0)
    g = lambda k: float(row[k] or 0)
    pre_conn = n('pre_total_conn')
    pre_talk_mins = g('pre_total_talk_mins')
    avg_talk_sec = (pre_talk_mins * 60.0 / pre_conn) if pre_conn > 0 else None
    return {
        'calls_attempted':  n('pre_total_calls'),
        'calls_connected':  pre_conn,
        'avg_talk_seconds': avg_talk_sec,
        'call_total_leads':    n('total_leads'),
        'call_pre_attempts':   n('pre_total_calls'),
        'call_pre_connects':   n('pre_total_conn'),
        'call_pre_talk_mins':  g('pre_total_talk_mins'),
        'call_pre_covered':    n('pre_covered_leads'),
        'call_p2_attempts':    n('p2_total_calls'),
        'call_p2_connects':    n('p2_total_conn'),
        'call_p2_talk_mins':   g('p2_total_talk_mins'),
        'call_p2_covered':     n('p2_covered_leads'),
        'call_p7_attempts':    n('p7_total_calls'),
        'call_p7_connects':    n('p7_total_conn'),
        'call_p7_talk_mins':   g('p7_total_talk_mins'),
        'call_p7_covered':     n('p7_covered_leads'),
        'call_p14_attempts':   n('p14_total_calls'),
        'call_p14_connects':   n('p14_total_conn'),
        'call_p14_talk_mins':  g('p14_total_talk_mins'),
        'call_p14_covered':    n('p14_covered_leads'),
        'call_p14p_attempts':  n('p14p_total_calls'),
        'call_p14p_connects':  n('p14p_total_conn'),
        'call_p14p_talk_mins': g('p14p_total_talk_mins'),
        'call_p14p_covered':   n('p14p_covered_leads'),
    }

def _query_sales_us(src_client, date_literal, wt_param):
    """US sales: count, USD revenue, YouTube-paid USD revenue.
    USD throughout (no INR conversion); paid-revenue filter is
    LOWER(Channel) = 'youtube' (covers any case spelling)."""
    from google.cloud import bigquery
    q = f"""
WITH sales AS (
  SELECT Sale_date, net_revenue, Channel, dupe_flag
  FROM (
    SELECT Sale_date, net_revenue, Channel, dupe_flag, webinar_type,
      DATE(event_start_date_time, "America/Los_Angeles") AS web_scheduled_date,
      ROW_NUMBER() OVER (
        PARTITION BY leads_hubspot_id, DATE(event_start_date_time, "America/Los_Angeles")
        ORDER BY formatted_date ASC
      ) AS rnk
    FROM `ik-marketing-data.Marketing_data_new_logic.Bq_data_Alumni`
    WHERE dupe_logic = 1
      AND COALESCE(Alumni_Stats, 'New_Lead') = 'New_Lead'   -- exclude alumni-repeat sales
  )
  WHERE rnk = 1
    AND web_scheduled_date IN ({date_literal})
    AND UPPER(webinar_type) = UPPER(@wt)
    AND dupe_flag = 0
)
SELECT
  COUNTIF(Sale_date IS NOT NULL)                                          AS sales,
  SUM(CASE WHEN Sale_date IS NOT NULL THEN net_revenue END)               AS revenue,
  -- Paid channels for US masterclass: L10X (paid masterclass funnel) +
  -- Google YouTube (YT ads). Matches the spend filter used upstream
  -- (LIKE '%l10x%' OR '%meta%' / '%facebook%'). Verified against
  -- Bq_data_Alumni distinct Channel values on 2026-06-22.
  SUM(CASE WHEN Sale_date IS NOT NULL AND Channel IN ('L10X', 'Google YouTube')
           THEN net_revenue END)                                          AS paid_revenue
FROM sales"""
    r = list(src_client.query(q, job_config=bigquery.QueryJobConfig(query_parameters=wt_param)).result())
    if not r:
        return {'sales': 0, 'revenue': 0.0, 'paid_revenue': 0.0}
    row = r[0]
    return {
        'sales':        int(row['sales'] or 0),
        'revenue':      float(row['revenue'] or 0.0),
        'paid_revenue': float(row['paid_revenue'] or 0.0),
    }

def _cumulative_snapshot_row(event_id, daily, live_at, now, now_iso,
                              cohort=None, attendance=None, call_data=None, email=None,
                              extra=None, sales_data=None):
    """Sum daily buckets to a single point-in-time snapshot row, merging in
    cohort + attendance + calls + email + extra (country-specific) metrics if available."""
    meta_regs   = sum(b['meta_regs']   for b in daily.values())
    meta_spend  = sum(b['meta_spend']  for b in daily.values())
    crm_regs    = sum(b['crm_regs']    for b in daily.values())
    other_regs  = sum(b['other_regs']  for b in daily.values())
    other_spend = sum(b['other_spend'] for b in daily.values())
    total_regs  = meta_regs + crm_regs + other_regs

    # Apply per-event spend correction (e.g., source pipeline dropped a day).
    correction = SPEND_CORRECTIONS.get(event_id, 0)
    if correction:
        meta_spend += correction
        print(f'  [snapshot/bq] {event_id}: applied spend correction +{correction:.0f}')

    hours_to_live = None
    if live_at is not None:
        live_at_aware = live_at if live_at.tzinfo else live_at.replace(tzinfo=datetime.timezone.utc)
        hours_to_live = (live_at_aware - now).total_seconds() / 3600.0

    row = {
        'event_id':         event_id,
        'snapshot_at':      now_iso,
        'hours_to_live':    hours_to_live,
        'total_regs':       int(total_regs),
        'meta_regs':        int(meta_regs),
        'meta_spend':       meta_spend,
        'crm_regs':         int(crm_regs),
        'other_regs':       int(other_regs),
        'other_spend':      other_spend,
        'cpiql':            (meta_spend / meta_regs) if meta_regs > 0 else None,
        'attendees': None, 'attendance_pct': None,
        'email_sent': None, 'email_delivered': None, 'email_opened': None, 'email_clicked': None,
        'calls_attempted': None, 'calls_connected': None, 'avg_talk_seconds': None,
        'role_sde': None, 'role_ml': None, 'role_fe': None,
        'role_management': None, 'role_systems': None, 'role_null': None, 'role_other': None,
        'we_0_2': None, 'we_3_5': None, 'we_6_10': None,
        'we_10_15': None, 'we_15_20': None, 'we_20p': None,
        'we_other': None, 'we_10p': None,
        'us_yt_regs': None, 'us_social_regs': None,
        'us_l10x_email_regs': None, 'us_l10x_bot_regs': None,
        'us_ni_base_regs': None, 'us_other_regs': None,
        'extras': None,
        'sales': None, 'revenue': None, 'paid_revenue': None,
        'overall_roas': None, 'paid_roas': None,
    }
    for partial in (cohort, attendance, call_data, email, extra):
        if partial:
            row.update(partial)
    if sales_data:
        row.update(sales_data)
        spend = row.get('meta_spend') or 0
        rev   = sales_data.get('revenue') or 0
        paid_rev = sales_data.get('paid_revenue') or 0
        row['overall_roas'] = round(rev / spend, 2) if spend > 0 else None
        row['paid_roas']    = round(paid_rev / spend, 2) if spend > 0 else None
    return row

def run_daily_snapshot():
    last_sent = None
    while True:
        now = datetime.datetime.now()
        if now.hour == 11 and now.minute < 2 and now.date() != last_sent:
            try:
                _do_snapshot()
                last_sent = now.date()
            except Exception as e:
                print(f'  [snapshot] error: {e}')
        time.sleep(60)

if __name__ == '__main__':
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    threading.Thread(target=run_daily_snapshot, daemon=True).start()
    print(f'Serving at http://localhost:{PORT}')
    print(f'Open: http://localhost:{PORT}/Masterclass%20Automation.html')
    http.server.ThreadingHTTPServer(('', PORT), Handler).serve_forever()
